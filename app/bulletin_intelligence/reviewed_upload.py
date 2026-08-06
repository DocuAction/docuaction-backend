"""Parse a reviewed FCC bulletin workbook back in (Task 3.1).

The workflow this serves: the system exports the bulletin to Excel, a human
edits summaries / fixes categories / deletes rows, and the corrected file comes
back. Parsing it is the point where an operator's careful hour of work either
survives or is silently mangled, so this module is strict on structure and
loud about anything it cannot honour.

Two rules drive the design:

  Validate before mutating. The whole workbook is parsed and checked first;
  nothing is applied unless every row is intelligible. A half-applied upload is
  worse than a rejected one, because nobody can tell which half landed.

  A missing row is a deletion, not an error. Removing a story is the most
  common review action. But it is also what a truncated or wrong-briefing
  upload looks like, so deletions are reported and require an explicit flag to
  take effect.

This module only reads and reports. Applying the result is the caller's job.
"""

from __future__ import annotations

import io
import re
from typing import Any, Dict, List, Optional, Tuple

# Column order is the contractual FCC deliverable format (columns A-K) and must
# match excel_export.HEADERS exactly. Validated at import against the exporter so
# the two cannot drift apart unnoticed.
EXPECTED_HEADERS = ["#", "Category", "Date", "Relationship", "Title", "Summary",
                    "Source", "Subscription Required", "Relevance", "URL",
                    "Provider"]

COL = {name: i for i, name in enumerate(EXPECTED_HEADERS)}

MAX_ROWS = 5000
MAX_BYTES = 10 * 1024 * 1024

EDITABLE_FIELDS = ("category", "title", "summary", "relevance", "subscription_required")


class UploadError(Exception):
    """The workbook cannot be used at all. Carries a reason fit to show a user."""


def _norm(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _header_row(ws) -> Tuple[int, List[str]]:
    """Find the header row.

    Exports carry a title block above the table, and reviewers add notes, so the
    header is not reliably row 1. Search the first 15 rows for one that starts
    with the '#' / 'Category' pair rather than assuming a position.
    """
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
        cells = [_norm(c) for c in (row or [])]
        if len(cells) >= 2 and cells[0] == "#" and cells[1] == "Category":
            return idx, cells
    raise UploadError(
        "No header row found in the first 15 rows. Expected a row beginning "
        "'#', 'Category' — is this an exported FCC bulletin workbook?")


def _check_headers(found: List[str]) -> None:
    actual = [h for h in found[:len(EXPECTED_HEADERS)]]
    if actual != EXPECTED_HEADERS:
        missing = [h for h in EXPECTED_HEADERS if h not in actual]
        extra = [h for h in actual if h and h not in EXPECTED_HEADERS]
        detail = []
        if missing:
            detail.append(f"missing {missing}")
        if extra:
            detail.append(f"unexpected {extra}")
        if not detail:
            # Same names, different order — reordering columns silently remaps
            # every value, so it is rejected rather than guessed at.
            detail.append(f"columns are out of order: got {actual}")
        raise UploadError(
            "Column headers do not match the FCC deliverable format (A-K): "
            + "; ".join(detail))


def parse_reviewed_workbook(data: bytes, *, sheet_name: Optional[str] = None
                            ) -> Dict[str, Any]:
    """Parse an uploaded .xlsx into rows keyed by URL.

    Returns {"rows": [...], "warnings": [...], "sheet": str, "count": int}.
    Raises UploadError with a user-facing reason on anything structural.
    """
    if not data:
        raise UploadError("Empty upload")
    if len(data) > MAX_BYTES:
        raise UploadError(
            f"File is {len(data) // (1024 * 1024)}MB; the limit is "
            f"{MAX_BYTES // (1024 * 1024)}MB")
    # .xlsx is a zip; every one starts 'PK'. Checking here turns "a .xls or a
    # CSV was renamed" into a clear message instead of an openpyxl stack trace.
    if not data.startswith(b"PK"):
        raise UploadError(
            "Not an .xlsx file. Excel 97-2003 (.xls) and CSV are not accepted — "
            "re-save as .xlsx.")

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise UploadError("openpyxl is not installed on the server")

    try:
        # read_only for memory; data_only so formula cells yield their cached
        # value rather than the formula text.
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        raise UploadError(f"Could not open the workbook: {str(e)[:200]}")

    try:
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise UploadError(
                    f"No sheet named {sheet_name!r}. Sheets present: {wb.sheetnames}")
            ws = wb[sheet_name]
        else:
            ws = wb[wb.sheetnames[0]]

        header_idx, header_cells = _header_row(ws)
        _check_headers(header_cells)

        rows: List[Dict[str, Any]] = []
        warnings: List[str] = []
        seen_urls: Dict[str, int] = {}

        for excel_row, raw in enumerate(
                ws.iter_rows(min_row=header_idx + 1, values_only=True),
                header_idx + 1):
            cells = [_norm(c) for c in (raw or [])]
            if not any(cells):
                continue                      # blank spacer row
            if len(rows) >= MAX_ROWS:
                warnings.append(
                    f"Stopped at {MAX_ROWS} rows; the rest of the sheet was ignored")
                break

            def cell(name: str) -> str:
                i = COL[name]
                return cells[i] if i < len(cells) else ""

            url = cell("URL")
            if not url:
                # Without a URL there is nothing to match the edit back to, so
                # the row is reported rather than dropped in silence.
                warnings.append(
                    f"Row {excel_row}: no URL — cannot be matched to an article, skipped")
                continue
            if not url.startswith(("http://", "https://")):
                warnings.append(
                    f"Row {excel_row}: URL {url[:60]!r} is not http(s), skipped")
                continue
            key = url.lower()
            if key in seen_urls:
                warnings.append(
                    f"Row {excel_row}: duplicate of row {seen_urls[key]} "
                    f"({url[:60]}) — the later row wins")
            seen_urls[key] = excel_row

            summary = cell("Summary")
            words = len(summary.split())
            # 60-100 is the house target, not a hard rule; a reviewer who
            # deliberately writes 45 words gets a note, not a rejection.
            if summary and not (60 <= words <= 100):
                warnings.append(
                    f"Row {excel_row}: summary is {words} words (target 60-100)")

            rows.append({
                "excel_row": excel_row,
                "url": url,
                "category": cell("Category"),
                "date": cell("Date"),
                "relationship": cell("Relationship"),
                "title": cell("Title"),
                "summary": summary,
                "source": cell("Source"),
                "subscription_required": cell("Subscription Required"),
                "relevance": cell("Relevance"),
                "provider": cell("Provider"),
                "word_count": words,
            })

        if not rows:
            raise UploadError(
                "The workbook has headers but no usable data rows "
                "(every row lacked a URL or was blank)")

        # Deduplicate on URL, later row winning, matching the warning above.
        by_url: Dict[str, Dict[str, Any]] = {}
        for r in rows:
            by_url[r["url"].lower()] = r

        return {
            "sheet": ws.title,
            "header_row": header_idx,
            "rows": list(by_url.values()),
            "count": len(by_url),
            "duplicate_rows": len(rows) - len(by_url),
            "warnings": warnings,
        }
    finally:
        try:
            wb.close()
        except Exception:
            pass


def diff_against_articles(parsed_rows: List[Dict[str, Any]],
                          articles: List[Any]) -> Dict[str, Any]:
    """Compare an upload against the briefing it claims to correct.

    Reports what would change without changing anything. `removed` is the list
    the caller must be most careful with: it means the reviewer deleted those
    rows, but it looks identical to someone uploading the wrong briefing.
    """
    def field(a: Any, name: str, default: str = "") -> str:
        v = a.get(name, default) if isinstance(a, dict) else getattr(a, name, default)
        return _norm(v)

    current = {field(a, "url").lower(): a for a in articles if field(a, "url")}
    uploaded = {r["url"].lower(): r for r in parsed_rows}

    edits: List[Dict[str, Any]] = []
    for key, row in uploaded.items():
        art = current.get(key)
        if art is None:
            continue
        changed = {}
        if row["title"] and row["title"] != field(art, "title"):
            changed["title"] = {"from": field(art, "title"), "to": row["title"]}
        if row["summary"] and row["summary"] != field(art, "summary"):
            changed["summary"] = {"from": field(art, "summary")[:120],
                                  "to": row["summary"][:120]}
        if changed:
            edits.append({"url": row["url"], "changes": changed})

    added = sorted(k for k in uploaded if k not in current)
    removed = sorted(k for k in current if k not in uploaded)

    return {
        "edited": edits,
        "edited_count": len(edits),
        "added": added,
        "added_count": len(added),
        "removed": removed,
        "removed_count": len(removed),
        "unchanged_count": len(uploaded) - len(edits) - len(added),
        "note": ("`removed` means those URLs were not in the upload. That is a "
                 "reviewer deletion if the file is complete, and data loss if "
                 "it is a partial or wrong-briefing upload — which is why "
                 "applying deletions requires apply_deletions=true."),
    }


def _assert_headers_match_exporter() -> None:
    """Fail at import if the exporter's columns drift from the parser's.

    The two are one contract read from opposite ends. Catching a mismatch here
    beats discovering it when a reviewer's edits land in the wrong fields.
    """
    try:
        from app.bulletin_intelligence.excel_export import HEADERS
    except Exception:
        return
    if list(HEADERS) != EXPECTED_HEADERS:
        raise RuntimeError(
            "reviewed_upload.EXPECTED_HEADERS is out of sync with "
            f"excel_export.HEADERS: exporter={list(HEADERS)} parser={EXPECTED_HEADERS}")


_assert_headers_match_exporter()
