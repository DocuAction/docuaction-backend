"""FCC bulletin -> client-facing Excel workbook (3 sheets).

Sheet 1 "FCC Daily Bulletin"      — one row per story, columns A-K
Sheet 2 "Google News Cross-Check" — stories Google News carried that our own
                                    sources did not, so a reviewer sees what the
                                    QA pass caught before the FCC does
Sheet 3 "Summary"                 — counts by category and relevance, coverage

Why articles are passed in rather than read off the briefing: the Briefing
dataclass stores `html_content`, `article_count` and `topic_counts` — it has no
article list. The caller rehydrates the stories by matching the URLs in the
stored HTML back to the archive, which is what keeps this workbook and the HTML
preview showing the same set rather than two independently re-derived ones.

COLUMNS H AND I ARE DELIVERABLE FIELDS, NOT INTERNAL SIGNALS. An earlier version
of this module excluded "Subscription Required" and "Relevance" on the theory
that they were internal QA data on a public endpoint. They are not: both are part
of the contract workbook's A-K column set. Subscription Required tells FCC staff
whether they can actually read the article; Relevance tells them what to
prioritise. Do not remove them as a "security tightening" — that would break the
deliverable format.
"""
from __future__ import annotations

import re
from html import unescape
from typing import Any, Callable, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NAVY = "003087"
BLUE = "0078D4"
BAND = "F5F8FD"
GRID = "D0D0D0"

HEADERS = ["#", "Category", "Date", "Relationship", "Title", "Summary", "Source",
           "Subscription Required", "Relevance", "URL", "Provider"]
WIDTHS = [5, 22, 15, 15, 50, 80, 25, 12, 10, 50, 15]
WRAP_COLS = (5, 6)      # Title, Summary
URL_COL = 10

_thin = Side(style="thin", color=GRID)
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


_TAG = re.compile(r"<[^>]+>")
_WS = re.compile(r"\s+")
_SPACE_BEFORE_PUNCT = re.compile(r"\s+([.,;:!?%)\]])")
# An UNTERMINATED tag at the end of the string. Summaries are truncated upstream,
# which routinely cuts mid-tag and leaves '... The post <a href="https://exa' with
# no closing '>'. _TAG cannot match that, so the raw markup reached the cell -- the
# exact defect observed on dev (16 cells, every one ending in <a/<img/<font).
# Requires a letter after '<' so ordinary prose like "5 < 10" is left alone.
_DANGLING_TAG = re.compile(r"<\s*/?[a-zA-Z][^>]*$")


def strip_html(text: Any) -> str:
    """Plain text from a value that may contain HTML.

    Summaries and titles are assembled from scraped feed content and from the
    briefing renderer, both of which carry markup. Written straight into a cell
    that markup is what the reader sees — "<p>The Commission voted…</p>" — which
    is the defect this exists to prevent.

    Order matters: tags are removed BEFORE entities are decoded. Decoding first
    would turn "&lt;script&gt;" into "<script>" and the tag stripper would then
    delete it, silently destroying text that was never markup to begin with.
    """
    if text is None:
        return ""
    s = _TAG.sub(" ", str(text))
    s = _DANGLING_TAG.sub(" ", s)
    s = unescape(s)
    # Re-strip: unescape can reveal a literal tag that was entity-encoded in the
    # source (&lt;div&gt;), which must not survive into the cell either.
    s = _TAG.sub(" ", s)
    s = _DANGLING_TAG.sub(" ", s)
    s = _WS.sub(" ", s).strip()
    # Tags are replaced with a space so "<b>a</b><b>b</b>" does not become "ab",
    # but that leaves "voted <b>today</b>." reading as "voted today ." — close the
    # gap before sentence punctuation.
    return _SPACE_BEFORE_PUNCT.sub(r"\1", s)


def _clean(value: Any) -> str:
    """The single choke point for every text cell in this workbook.

    Strips HTML, decodes entities, collapses whitespace, and drops control
    characters — Excel rejects the latter and openpyxl raises on them, so one bad
    scraped summary would otherwise cost the whole download.
    """
    s = strip_html(value)
    return "".join(ch for ch in s if ord(ch) >= 32).strip()


def _sheet_title(prefix: str, date_str: str) -> str:
    """Excel sheet names cap at 31 chars and forbid []:*?/\\ ."""
    raw = f"{prefix} {date_str}".strip()
    for bad in "[]:*?/\\":
        raw = raw.replace(bad, "")
    return raw[:31] or "Bulletin"


def _source_of(article: Any) -> str:
    """Prefer the human outlet name; fall back through the collector fields.

    Routed through _clean like every other text cell — outlet names arrive from
    feed metadata and have carried entities (&amp;) before now.
    """
    for attr in ("source_name", "outlet", "source"):
        v = getattr(article, attr, "") or ""
        if v:
            return _clean(v)
    return ""


# Internal-only columns appended to the QA workbook after the shared A-K set.
QA_HEADERS = ["QA Score", "Duplicate Flag", "URL Status", "Word Count",
              "Google News Match", "Held From Bulletin"]
QA_WIDTHS = [10, 14, 12, 11, 18, 20]


def _relevance_band(article: Any) -> str:
    try:
        score = float(getattr(article, "relevance_score", 0) or 0)
    except (TypeError, ValueError):
        return "Low"
    if score >= 0.75:
        return "High"
    return "Medium" if score >= 0.45 else "Low"


def _relationship(article: Any) -> str:
    """Original / Follow-up / Analysis, derived from the classifier's type."""
    kind = (getattr(article, "article_type", "") or "").lower()
    if kind in ("analysis", "opinion", "editorial"):
        return "Analysis"
    if kind in ("follow-up", "followup"):
        return "Follow-up"
    return "Original"


def _title_with_tags(article: Any) -> str:
    """Title carrying the [Opinion] / [Subscription Required] tags.

    The tags belong in the title by design — the summary prompt forbids them in
    the summary, so this is the only place a reader sees them.
    """
    title = _clean(getattr(article, "title", ""))
    kind = (getattr(article, "article_type", "") or "").lower()
    if kind in ("opinion", "editorial") and "[Opinion]" not in title:
        title = f"{title} [Opinion]"
    if bool(getattr(article, "is_paywalled", False)) and "[Subscription Required]" not in title:
        title = f"{title} [Subscription Required]"
    return title


def _date_of(article: Any) -> str:
    return _clean(getattr(article, "published_at", ""))[:10]


def _style_header(ws, headers: List[str], row: int = 1) -> None:
    font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor=NAVY)
    for col, name in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=name)
        c.font = font
        c.fill = fill
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        widths = WIDTHS + QA_WIDTHS
        if col - 1 < len(widths):
            ws.column_dimensions[get_column_letter(col)].width = widths[col - 1]


def _write_rows(ws, rows: List[Any], topic_of, start_row: int = 2,
                extras: Optional[Callable[[Any], List[Any]]] = None) -> int:
    """Write the A-K body, plus `extras(article)` appended when supplied.

    The QA workbook is defined as "the client workbook plus internal columns", so
    it shares this writer rather than maintaining a parallel column list — that is
    what keeps A-K identical between the two files instead of drifting apart.
    """
    body = Font(name="Arial", size=10)
    cat_font = Font(name="Arial", size=10, bold=True, color=NAVY)
    url_font = Font(name="Arial", size=10, color=BLUE, underline="single")
    band_fill = PatternFill("solid", fgColor=BAND)
    wrap = Alignment(wrap_text=True, vertical="top")
    top = Alignment(vertical="top")

    r = start_row - 1
    for i, a in enumerate(rows, start=1):
        r = start_row + i - 1
        url = _clean(getattr(a, "url", ""))
        values = [
            i,
            topic_of(a),
            _date_of(a),
            _relationship(a),
            _title_with_tags(a),
            _clean(getattr(a, "summary", "")),
            _source_of(a),
            "Yes" if getattr(a, "is_paywalled", False) else "No",
            _relevance_band(a),
            url,
            _clean(getattr(a, "provider", "") or getattr(a, "source", "")),
        ]
        if extras is not None:
            try:
                values = values + list(extras(a))
            except Exception:  # noqa: BLE001 — a QA column must not cost the file
                values = values + [""] * len(QA_HEADERS)
        for col, v in enumerate(values, start=1):
            c = ws.cell(row=r, column=col, value=v)
            c.font = body
            c.border = BORDER
            c.alignment = wrap if col in WRAP_COLS else top
            if i % 2 == 0:
                c.fill = band_fill
        ws.cell(row=r, column=2).font = cat_font
        u = ws.cell(row=r, column=URL_COL)
        u.font = url_font
        # Excel refuses very long hyperlink targets, and one bad link should not
        # cost the whole workbook.
        if url.startswith(("http://", "https://")) and len(url) <= 2000:
            u.hyperlink = url
    return r


def create_bulletin_excel(
    briefing: Dict[str, Any],
    articles: List[Any],
    section_of: Optional[Callable[[Any], str]] = None,
    google_news_missing: Optional[List[Any]] = None,
    qa_report: Optional[Dict[str, Any]] = None,
    qa: bool = False,
    qa_extras: Optional[Callable[[Any], List[Any]]] = None,
) -> Workbook:
    """Build the three-sheet client workbook.

    `section_of` maps an article to its display section (the engine's
    `_section_of`). When absent we fall back to the article's own `section` or
    `topic`, so this module never hard-depends on the engine.

    `google_news_missing` are stories Google News carried that our own sources
    did not. Empty or omitted renders Sheet 2 as an explicit "all matched"
    statement rather than a blank sheet — a blank sheet reads as "not run", and
    this one needs to say "ran, found nothing".
    """
    date_str = _clean(briefing.get("briefing_date") or "")
    agency = str(briefing.get("agency_id") or "fcc").upper()

    def topic_of(a: Any) -> str:
        if section_of is not None:
            try:
                return _clean(section_of(a)) or "General"
            except Exception:
                pass
        return _clean(getattr(a, "section", "") or getattr(a, "topic", "") or "General")

    # Sort by category, then source within category — skimmable by section.
    rows = sorted(articles, key=lambda a: (topic_of(a).lower(), _source_of(a).lower()))

    wb = Workbook()

    # QA mode is the same workbook plus internal columns L-P. Sharing one path is
    # deliberate: the two downloads must agree on A-K, and a second column list
    # would drift the moment either is edited.
    headers = HEADERS + QA_HEADERS if qa else HEADERS
    extras = qa_extras if qa else None

    # ── Sheet 1: FCC Daily Bulletin ───────────────────────────────────────────
    ws = wb.active
    ws.title = "FCC Daily Bulletin"
    _style_header(ws, headers)
    last = _write_rows(ws, rows, topic_of, extras=extras)
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last}"

    # ── Sheet 2: Google News Cross-Check ──────────────────────────────────────
    gn = wb.create_sheet("Google News Cross-Check")
    missing = list(google_news_missing or [])

    banner = gn.cell(row=1, column=1, value=(
        "GOOGLE NEWS CROSS-CHECK — Articles found in Google News but not in "
        "primary sources. Review for potential inclusion."))
    banner.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    banner.fill = PatternFill("solid", fgColor=NAVY)
    banner.alignment = Alignment(horizontal="left", vertical="center")
    gn.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    if missing:
        _style_header(gn, headers, row=2)
        gn_last = _write_rows(gn, missing, topic_of, start_row=3, extras=extras)
        gn.freeze_panes = "A3"
        gn.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{gn_last}"
    else:
        msg = gn.cell(row=3, column=1, value=(
            "All Google News articles matched. No missing stories identified."))
        msg.font = Font(name="Arial", size=11, color=NAVY)
        for col in range(1, len(headers) + 1):
            gn.column_dimensions[get_column_letter(col)].width = (WIDTHS + QA_WIDTHS)[col - 1]

    # ── Sheet 3: Summary ──────────────────────────────────────────────────────
    s = wb.create_sheet("Summary")
    s.column_dimensions["A"].width = 42
    s.column_dimensions["B"].width = 16

    s["A1"] = f"{agency} Daily Intelligence Bulletin"
    s["A1"].font = Font(name="Arial", size=14, bold=True, color=NAVY)
    s["A2"] = date_str
    s["A2"].font = Font(name="Arial", size=12, color=NAVY)
    s["A3"] = "Prepared by Alliance Global Tech, Inc. (AGT)"
    s["A3"].font = Font(name="Arial", size=10)

    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=NAVY)
    body = Font(name="Arial", size=10)
    label_font = Font(name="Arial", size=10, bold=True, color=NAVY)

    def _stat(row: int, label: str, value: Any, bold: bool = False) -> int:
        lc = s.cell(row=row, column=1, value=label)
        lc.font = label_font if bold else body
        lc.border = BORDER
        vc = s.cell(row=row, column=2, value=value)
        vc.font = label_font if bold else body
        vc.border = BORDER
        vc.alignment = Alignment(horizontal="center")
        return row + 1

    for col, name in enumerate(("Statistic", "Value"), start=1):
        c = s.cell(row=5, column=col, value=name)
        c.font = header_font
        c.fill = header_fill
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")

    r = 6
    r = _stat(r, "Total Articles", len(rows), bold=True)

    # Counts come from the rows actually written, so the sheets can never
    # disagree — deriving them from briefing["topic_counts"] would reintroduce
    # the count-vs-render split this workbook exists to avoid.
    counts: Dict[str, int] = {}
    for a in rows:
        key = topic_of(a)
        counts[key] = counts.get(key, 0) + 1
    ordered = sorted(counts.items(), key=lambda kv: (-kv[1], kv[0].lower()))

    r += 1
    s.cell(row=r, column=1, value="By Category").font = label_font
    r += 1
    cat_start = r
    for cat, n in ordered:
        r = _stat(r, cat, n)
    cat_end = r - 1

    bands = {"High": 0, "Medium": 0, "Low": 0}
    for a in rows:
        bands[_relevance_band(a)] += 1
    r += 1
    s.cell(row=r, column=1, value="By Relevance").font = label_font
    r += 1
    for band in ("High", "Medium", "Low"):
        r = _stat(r, band, bands[band])

    rep = qa_report or {}
    total_gn = int(rep.get("google_news_count", 0) or 0)
    matched = int(rep.get("matched", 0) or 0)
    rate = f"{(matched / total_gn * 100):.0f}%" if total_gn else "n/a"
    r += 1
    r = _stat(r, "Google News Coverage", f"{matched}/{total_gn} ({rate})")
    r = _stat(r, "Missing from Google News", len(missing))
    r = _stat(r, "Sources Used", len({_source_of(a) for a in rows if _source_of(a)}))

    r += 1
    tc = s.cell(row=r, column=1, value="TOTAL")
    tc.font = Font(name="Arial", size=11, bold=True, color=NAVY)
    tc.border = BORDER
    nc = s.cell(row=r, column=2)
    # A live formula, not a baked number, so the sheet stays correct if a reader
    # deletes a category row.
    nc.value = f"=SUM(B{cat_start}:B{cat_end})" if ordered else 0
    nc.font = Font(name="Arial", size=11, bold=True, color=NAVY)
    nc.border = BORDER
    nc.alignment = Alignment(horizontal="center")

    return wb
