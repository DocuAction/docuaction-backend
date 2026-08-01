"""Export endpoints - download data as Excel files."""
import io
from fastapi import APIRouter, Depends
from fastapi.responses import Response
from sqlalchemy import select, desc
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from app.core.security import require_role
from app.database import get_db
from app.models import RFQ, Quote, Candidate, Application, JobPosting
from app.services.auth import get_current_user

# Router-level auth. app/routers/ is dormant (see __init__.py) and this
# dependency is the precondition recorded there for ever mounting it: every
# route inherits the check, so a handler added later cannot arrive unguarded.
router = APIRouter(prefix="/export", tags=["Export"], dependencies=[Depends(require_role("contributor"))])
def _make_xlsx(headers, rows, sheet_name="Export"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header styling
    header_fill = PatternFill(start_color="0A1628", end_color="0A1628", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    border = Border(
        bottom=Side(style='thin', color='DEE2E6'),
        right=Side(style='thin', color='DEE2E6'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical='top')

    # Auto-width
    for col in ws.columns:
        max_len = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@router.get("/rfqs")
async def export_rfqs(user=Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(RFQ).order_by(desc(RFQ.created_at)))
    rfqs = result.scalars().all()
    headers = ["Title", "Solicitation #", "Agency", "Status", "Type", "Due Date", "Value", "Officer Name", "Officer Email", "Officer Phone", "Created"]
    rows = []
    for r in rfqs:
        rows.append([
            r.title, r.solicitation_number or '', r.agency or '', str(r.status),
            str(r.customer_type), str(r.due_date) if r.due_date else '',
            float(r.estimated_value) if r.estimated_value else 0,
            getattr(r, 'contract_officer_name', '') or '',
            getattr(r, 'contract_officer_email', '') or '',
            getattr(r, 'contract_officer_phone', '') or '',
            str(r.created_at)[:10] if r.created_at else '',
        ])
    xlsx = _make_xlsx(headers, rows, "RFQs")
    return Response(content=xlsx, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": 'attachment; filename="AGT_RFQs_Export.xlsx"'})


@router.get("/quotes")
async def export_quotes(user=Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Quote).order_by(desc(Quote.created_at)))
    quotes = result.scalars().all()
    headers = ["Quote #", "Version", "Status", "Sell Price", "Cost", "Margin %", "Tax", "Created"]
    rows = []
    for q in quotes:
        rows.append([
            q.quote_number or '', q.version, str(q.status),
            float(q.total_sell_price or 0), float(q.total_cost or 0),
            float(q.overall_margin_pct or 0), float(q.total_tax or 0),
            str(q.created_at)[:10] if q.created_at else '',
        ])
    xlsx = _make_xlsx(headers, rows, "Quotes")
    return Response(content=xlsx, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": 'attachment; filename="AGT_Quotes_Export.xlsx"'})


@router.get("/candidates")
async def export_candidates(user=Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Candidate).order_by(desc(Candidate.created_at)))
    cands = result.scalars().all()
    headers = ["First Name", "Last Name", "Email", "Phone", "Location", "Skills", "Experience", "Clearance", "Source", "Created"]
    rows = []
    for c in cands:
        rows.append([
            c.first_name, c.last_name, c.email, c.phone or '',
            c.location or '', c.skills or '',
            c.years_experience or '', c.clearance_level or '',
            c.source or '', str(c.created_at)[:10] if c.created_at else '',
        ])
    xlsx = _make_xlsx(headers, rows, "Candidates")
    return Response(content=xlsx, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": 'attachment; filename="AGT_Candidates_Export.xlsx"'})
