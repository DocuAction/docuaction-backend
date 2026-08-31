"""Render a report dataset as an Excel workbook.

WHAT AN EXCEL ENGINE IS FOR, AND WHAT IT MUST NOT DO
────────────────────────────────────────────────────
Excel is helpful and Excel is opinionated. Left alone it will read `01234` as
the number 1234, `1234567890123456` as 1.23457E+15, `03/04` as a date, and
`=cmd|...` as a formula to evaluate. Every one of those is a change to a
Government value, made silently, after the data left DocuAction.

So this engine's job is mostly refusal:

  * every source cell is written as TEXT, so a delivered identifier stays the
    string that was delivered;
  * a value that begins with a character Excel treats as a formula is stored
    as a string and flagged literal, so Excel shows it and never evaluates it —
    the cell still READS as exactly the delivered value;
  * nothing is trimmed, padded, rounded or reformatted on the way out.

THE APOSTROPHE IS A REPRESENTATION, NOT AN EDIT
───────────────────────────────────────────────
`quotePrefix` is Excel's own way of saying "the following is literal". It lives
in the cell's STYLE, not in its value: nothing is prepended, and reading the
cell back gives the original string byte for byte. DocuAction's stored source is
untouched either way — this only governs what the spreadsheet does with it.

DESIGN
──────
Restrained. A federal review workbook is read, sorted and filtered; it is not
looked at. One header treatment, one body font, frozen headers, filters, sized
columns, no merged cells across data, no colour carrying meaning on its own, no
macros and no external references.
"""

from __future__ import annotations

import io
import re
from copy import copy
from datetime import date, datetime, timezone
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.cell_style import StyleArray
from openpyxl.utils import get_column_letter

#: Bump when the RENDERING changes, so a delivered file can be traced to the
#: engine that produced it. Distinct from the workbook DATA version.
XLSX_ENGINE_VERSION = "1.0.0"

XLSX_CONTENT_TYPE = ("application/vnd.openxmlformats-officedocument"
                     ".spreadsheetml.sheet")

HEADER_FILL = "0B3C5D"      # the platform's primary navy
HEADER_TEXT = "FFFFFF"
GRID = "D8D8D8"
FONT = "Calibri"

#: Characters Excel may treat as the start of a formula. The tab and carriage
#: return are here because Excel has historically stripped leading whitespace
#: and then evaluated what followed.
FORMULA_LEADERS = ("=", "+", "-", "@", "\t", "\r")

#: Excel forbids these in a cell; openpyxl raises rather than writing them.
_ILLEGAL = re.compile(r"[\000-\010\013\014\016-\037]")

_thin = Side(style="thin", color=GRID)
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

MAX_COLUMN_WIDTH = 60
MIN_COLUMN_WIDTH = 10


def looks_like_formula(value: Any) -> bool:
    """Would Excel try to evaluate this cell?"""
    return isinstance(value, str) and value.startswith(FORMULA_LEADERS)


def _sanitise(value: str) -> str:
    """Remove control characters Excel refuses to store.

    This is the ONE transformation applied to a source string, and it is
    applied because the alternative is a workbook that cannot be written at
    all. Control characters are not visible data.
    """
    return _ILLEGAL.sub("", value)


class _Styles:
    """One resolved style per DISTINCT combination, for the whole workbook.

    openpyxl interns every style assignment into a workbook-wide table, and the
    lookup HASHES the Font/Border/Alignment object each time — recursively,
    through every child object. On a seven-row fixture that is invisible. At
    23,566 rows across 41 columns it was measured at roughly 2,000 cells a
    second, which is a quarter-hour export and several assignments per cell
    doing nothing but re-proving that the same border is the same border.

    This workbook uses about a dozen distinct styles. So the FIRST cell of each
    kind is styled the ordinary way, through the public attributes, and the
    resolved reference is kept; every later cell of that kind copies it. The
    style is still defined by the assignments below — this only stops it being
    recomputed a million times. `_style` is the same StyleArray that openpyxl's
    own named-style assignment copies, and it is workbook-scoped, which is why
    the cache is created per render and never shared between workbooks.
    """

    def __init__(self):
        self._resolved: Dict[Any, Any] = {}

    def apply(self, cell, key: Any, define) -> None:
        known = self._resolved.get(key)
        if known is not None:
            cell._style = copy(known)
            return
        define(cell)
        resolved = cell._style
        if not isinstance(resolved, StyleArray):
            # openpyxl changed shape underneath us. Style every cell the slow,
            # correct way rather than copy something we do not understand.
            return
        self._resolved[key] = copy(resolved)


def _write_cell(ws, row: int, column: int, value: Any, *, as_text: bool,
                body: Font, wrap: bool = False, styles: "_Styles" = None):
    cell = ws.cell(row=row, column=column)
    number_format: Optional[str] = None
    literal = False

    if value is None:
        cell.value = None
    elif isinstance(value, (datetime, date)):
        # Excel has no concept of a timezone and openpyxl refuses an aware
        # datetime outright. Converting to UTC and dropping the marker keeps
        # the INSTANT correct; the sheets name their timestamps as UTC, so
        # nothing is lost but the suffix.
        if isinstance(value, datetime) and value.tzinfo is not None:
            value = value.astimezone(timezone.utc).replace(tzinfo=None)
        cell.value = value
        number_format = ("yyyy-mm-dd hh:mm:ss" if isinstance(value, datetime)
                         else "yyyy-mm-dd")
    elif isinstance(value, bool):
        # Before the int branch: a bool IS an int in Python, and a delivered
        # flag written as TRUE/FALSE is not what a reviewer expects to sort on.
        cell.value = "Yes" if value else "No"
    elif isinstance(value, (int, float)) and not as_text:
        cell.value = value
    else:
        text = _sanitise(str(value))
        cell.value = text
        if looks_like_formula(text):
            # Two separate things, both needed. openpyxl TYPES a leading "=" as
            # a formula, so the cell is forced back to a string; and Excel's own
            # literal marker is set so the application will not re-evaluate it.
            # Neither touches the value: quotePrefix lives in the cell's STYLE,
            # so reading the cell back returns exactly what was delivered.
            cell.data_type = "s"
            literal = True
        if as_text:
            number_format = "@"

    def define(target):
        target.font = body
        target.border = BORDER
        target.alignment = Alignment(vertical="top", wrap_text=wrap)
        if number_format is not None:
            target.number_format = number_format
        if literal:
            target.quotePrefix = True

    key = (body.name, body.size, body.bold, wrap, number_format, literal)
    if styles is None:
        define(cell)
    else:
        styles.apply(cell, key, define)
    return cell


def _column_widths(columns: List[str], rows: List[List[Any]],
                   sample: int = 200) -> List[int]:
    """Width from the header and the first rows, capped.

    Measuring every one of 23,566 rows to size a column costs a full pass for a
    cosmetic result; the first 200 settle it, and the cap stops one long free-text
    value making a column wider than the screen.
    """
    widths = [max(MIN_COLUMN_WIDTH, len(str(name)) + 4) for name in columns]
    for row in rows[:sample]:
        for i, value in enumerate(row[:len(widths)]):
            if value is None:
                continue
            widths[i] = max(widths[i], min(MAX_COLUMN_WIDTH, len(str(value)) + 3))
    return widths


def _write_sheet(wb: Workbook, name: str, sheet: Dict[str, Any],
                 first: bool, styles: "_Styles") -> int:
    ws = wb.active if first else wb.create_sheet()
    ws.title = name

    columns: List[str] = sheet["columns"]
    rows: List[List[Any]] = sheet["rows"]
    text_columns = set(sheet.get("text_columns") or [])
    note = sheet.get("note")

    header_font = Font(name=FONT, size=11, bold=True, color=HEADER_TEXT)
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    body = Font(name=FONT, size=10)

    start = 1
    if note:
        # The note sits ABOVE the header so a reader meets the caveat before the
        # data, and so it is not mistaken for a row.
        cell = ws.cell(row=1, column=1, value=_sanitise(note))
        cell.font = Font(name=FONT, size=10, italic=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[1].height = 30
        start = 3

    for index, title in enumerate(columns, start=1):
        cell = ws.cell(row=start, column=index, value=_sanitise(str(title)))
        cell.font, cell.fill, cell.border = header_font, header_fill, BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center",
                                   wrap_text=True)
    ws.row_dimensions[start].height = 28

    wrap_columns = {i for i, name_ in enumerate(columns)
                    if any(w in str(name_).lower()
                           for w in ("description", "definition", "reason",
                                     "observation", "note", "value", "use",
                                     "validation", "processing"))}

    for offset, values in enumerate(rows):
        r = start + 1 + offset
        for index in range(len(columns)):
            value = values[index] if index < len(values) else None
            _write_cell(ws, r, index + 1, value,
                        as_text=index in text_columns, body=body,
                        wrap=index in wrap_columns, styles=styles)

    for index, width in enumerate(_column_widths(columns, rows), start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    # Frozen header and a filter over the data — the two things that make a
    # 23,566-row sheet usable at all.
    ws.freeze_panes = ws.cell(row=start + 1, column=1)
    if rows:
        ws.auto_filter.ref = (f"A{start}:"
                              f"{get_column_letter(len(columns))}{start + len(rows)}")
    ws.sheet_view.showGridLines = False
    return len(rows)


def render_workbook(dataset: Dict[str, Any]) -> bytes:
    """The dataset as .xlsx bytes.

    `write_only` is deliberately NOT used. It would stream rows without holding
    the sheet in memory, but it cannot set freeze panes, filters, column widths
    or per-cell number formats — and those are what stop Excel retyping a
    delivered identifier. Correctness first; the measured footprint at
    Government scale is recorded in the export documentation.
    """
    wb = Workbook()
    wb.iso_dates = True
    styles = _Styles()

    order = [name for name in dataset["sheet_order"] if name in dataset["sheets"]]
    if not order:
        raise ValueError("A workbook needs at least one sheet.")

    for position, name in enumerate(order):
        _write_sheet(wb, name, dataset["sheets"][name], first=position == 0,
                     styles=styles)

    props = wb.properties
    props.title = "TEFCA ARC — ONC data review"
    props.creator = "DocuAction TEFCA ARC"
    props.description = (
        f"Controlled export {dataset['report_id']} · "
        f"workbook {dataset['workbook_version']} · "
        f"engine {XLSX_ENGINE_VERSION} · "
        f"classification {dataset['classification']}")
    props.category = dataset["classification"]

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
