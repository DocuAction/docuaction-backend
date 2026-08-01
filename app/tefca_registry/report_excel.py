"""Weekly report as a three-sheet Excel workbook.

Built from the ARCHIVED report_data, never recomputed. A report that renders one
set of numbers as HTML and a different set as Excel is worse than having no
Excel export at all, so both read the same stored snapshot.

Sheet 3 (Limitations) is not optional. If a reader opens the Excel and not the
HTML, they must still see what could not be checked — otherwise the export
quietly launders a caveated report into an uncaveated spreadsheet.
"""
from __future__ import annotations

import io
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NAVY = "003087"
BAND = "F5F8FD"
GRID = "D0D0D0"
FONT = "Arial"

_thin = Side(style="thin", color=GRID)
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

ENTITY_HEADERS = ["Review ID", "Entity", "NPI", "Type", "NPPES", "PECOS",
                  "OIG LEIE", "Bucket", "Rule", "Rationale"]
ENTITY_WIDTHS = [18, 38, 14, 16, 13, 13, 13, 9, 12, 70]


def _clean(v: Any) -> str:
    """Excel rejects control characters; strip rather than fail the download."""
    s = "" if v is None else str(v)
    return "".join(c for c in s if c in "\n\t" or ord(c) >= 32).strip()


def _header(ws, headers: List[str], widths: List[int], row: int = 1) -> None:
    f = Font(name=FONT, size=11, bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor=NAVY)
    for col, name in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=name)
        c.font, c.fill, c.border = f, fill, BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
        if col <= len(widths):
            ws.column_dimensions[get_column_letter(col)].width = widths[col - 1]


def _write_rows(ws, rows: List[List[Any]], start: int = 2,
                wrap_cols: tuple = ()) -> None:
    body = Font(name=FONT, size=10)
    band = PatternFill("solid", fgColor=BAND)
    wrap = Alignment(wrap_text=True, vertical="top")
    top = Alignment(vertical="top")
    for i, values in enumerate(rows):
        r = start + i
        for col, v in enumerate(values, start=1):
            c = ws.cell(row=r, column=col, value=_clean(v) if isinstance(v, str) else v)
            c.font, c.border = body, BORDER
            c.alignment = wrap if col in wrap_cols else top
            if i % 2:
                c.fill = band


def build_weekly_excel(report_data: Dict[str, Any], report_id: str,
                       entity_rows: List[Dict[str, Any]] | None = None) -> bytes:
    """Three sheets: entity results, summary statistics, limitations."""
    d = report_data or {}
    wb = Workbook()

    # ── Sheet 1: entity results ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Entity Results"
    _header(ws, ENTITY_HEADERS, ENTITY_WIDTHS)

    rows: List[List[Any]] = []
    for e in (entity_rows or []):
        v = e.get("verification") or {}
        rows.append([
            e.get("review_id"), e.get("entity_name"), e.get("npi"),
            e.get("entity_type"),
            (v.get("nppes") or {}).get("status"),
            (v.get("pecos") or {}).get("status"),
            (v.get("oig_leie") or {}).get("status"),
            e.get("bucket"), e.get("rule_code"), e.get("rationale"),
        ])
    if not rows:
        # An empty sheet with only headers reads as a failed export. Say why.
        rows = [["—"] * len(ENTITY_HEADERS)]
        rows[0][1] = "No reviews were completed in this period"
    _write_rows(ws, rows, wrap_cols=(10,))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(ENTITY_HEADERS))}{len(rows) + 1}"

    # ── Sheet 2: summary statistics ──────────────────────────────────────────
    s2 = wb.create_sheet("Summary Statistics")
    s2["A1"] = f"TEFCA ARC — {d.get('report_type', 'weekly').title()} Report"
    s2["A1"].font = Font(name=FONT, size=14, bold=True, color=NAVY)
    s2["A2"] = f"{report_id}   |   Contract {d.get('contract', '')}"
    s2["A2"].font = Font(name=FONT, size=11, color=NAVY)
    period = d.get("period") or {}
    s2["A3"] = f"Period {period.get('start', '')} to {period.get('end', '')}"
    s2["A3"].font = Font(name=FONT, size=10)

    _header(s2, ["Measure", "Value"], [46, 30], row=5)
    counts = (d.get("classification_distribution") or {}).get("counts") or {}
    ci = d.get("discrepancy_rate") or {}
    ex = d.get("executive_summary") or {}
    cfg = d.get("configuration") or {}
    stat_rows = [
        ["Entities reviewed", ex.get("entities_reviewed", 0)],
        ["B1 — No Discrepancy", counts.get("B1", 0)],
        ["B2 — Minor / Administrative", counts.get("B2", 0)],
        ["B3 — Inexplicable (manual review)", counts.get("B3", 0)],
        ["B4 — Non-Compliant", counts.get("B4", 0)],
        ["Discrepancies (anything not B1)", ex.get("discrepancies_found", 0)],
        ["Discrepancy rate", ci.get("rate") if ci.get("rate") is not None else "n/a"],
        ["Confidence interval (lower)", ci.get("lower", "n/a")],
        ["Confidence interval (upper)", ci.get("upper", "n/a")],
        ["Interval method", ci.get("method", "n/a")],
        ["B3 pending manual review", ex.get("b3_pending_manual_review", 0)],
        ["B4 requiring action", ex.get("b4_requiring_action", 0)],
        ["Rule set version", cfg.get("rule_set_version")],
        ["Confidence level", cfg.get("confidence_level")],
        ["Margin of error", cfg.get("margin_of_error")],
        ["Proportion", cfg.get("proportion")],
        ["Finite population correction", cfg.get("use_fpc")],
        ["Random seed", cfg.get("random_seed")],
        ["Generated at", cfg.get("generated_at")],
    ]
    _write_rows(s2, stat_rows, start=6)

    # Verification coverage, per source, keeping the five states distinct.
    cov = d.get("verification_coverage") or {}
    if cov:
        base = 6 + len(stat_rows) + 2
        s2.cell(row=base - 1, column=1, value="Verification coverage by source").font = \
            Font(name=FONT, size=12, bold=True, color=NAVY)
        _header(s2, ["Source", "Verified", "Not found", "Unavailable",
                     "Not checked", "Failed"],
                [26, 12, 12, 13, 13, 10], row=base)
        _write_rows(s2, [[src, r.get("verified", 0), r.get("not_found", 0),
                          r.get("unavailable", 0), r.get("not_checked", 0),
                          r.get("failed", 0)]
                         for src, r in sorted(cov.items())], start=base + 1)

    # Quarterly only: the per-week series.
    trend = d.get("weekly_trend")
    if trend:
        s4 = wb.create_sheet("Weekly Trend")
        _header(s4, ["Week", "B1", "B2", "B3", "B4"], [16, 10, 10, 10, 10])
        _write_rows(s4, [[t["week"], t["b1"], t["b2"], t["b3"], t["b4"]]
                         for t in trend])
        s4.freeze_panes = "A2"

    # ── Sheet 3: limitations (mandatory) ─────────────────────────────────────
    s3 = wb.create_sheet("Limitations")
    s3["A1"] = "Limitations and Exceptions"
    s3["A1"].font = Font(name=FONT, size=14, bold=True, color=NAVY)
    s3["A2"] = ("Always present. A report that omits what could not be checked "
                "invites the reader to assume full coverage.")
    s3["A2"].font = Font(name=FONT, size=10, italic=True)
    s3.column_dimensions["A"].width = 120
    _header(s3, ["Limitation"], [120], row=4)
    lims = d.get("limitations") or ["None identified."]
    _write_rows(s3, [[x] for x in lims], start=5, wrap_cols=(1,))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
