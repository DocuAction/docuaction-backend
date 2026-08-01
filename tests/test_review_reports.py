"""Report generation: completeness, arithmetic, and the mandatory limitations.

The load-bearing assertions here are the ones about honesty rather than format:
that limitations are never omitted, that an unreachable source is not reported
as a discrepancy, and that an empty period says so instead of printing zeros
that read like a clean bill of health.
"""
from datetime import date

import pytest

from app.tefca_registry.bucket_classifier import (
    FAILED, NOT_CHECKED, NOT_FOUND, UNAVAILABLE, VERIFIED)
from app.tefca_registry.report_generator import (
    BUCKET_LABELS, build_limitations, build_report_data, render_html,
    report_id_for)

START, END = date(2026, 7, 27), date(2026, 8, 2)

REQUIRED_SECTIONS = [
    "executive_summary", "sampling_summary", "classification_distribution",
    "discrepancy_rate", "verification_coverage", "outstanding_items",
    "data_sources_used", "methodology", "limitations", "configuration",
]


def review(rid, bucket, resolution=None, reclass=None):
    return {"review_id": rid, "classification_bucket": bucket,
            "reviewer_resolution": resolution, "reclassified_to": reclass}


def verif(source, status, n=1):
    return [{"source": source, "verification_status": status} for _ in range(n)]


def sample_reviews():
    return [review("REV-2026-000001", "B1"), review("REV-2026-000002", "B1"),
            review("REV-2026-000003", "B2"), review("REV-2026-000004", "B3"),
            review("REV-2026-000005", "B4")]


def sample_verifications():
    return (verif("nppes", VERIFIED, 4) + verif("nppes", NOT_FOUND, 1)
            + verif("pecos", VERIFIED, 3) + verif("pecos", UNAVAILABLE, 2)
            + verif("oig_leie", VERIFIED, 5))


def build(reviews=None, verifications=None, **kw):
    return build_report_data(
        report_type="weekly", period_start=START, period_end=END,
        reviews=sample_reviews() if reviews is None else reviews,
        verifications=sample_verifications() if verifications is None else verifications,
        rule_set_version=1, **kw)


# ── completeness ─────────────────────────────────────────────────────────────

@pytest.mark.parametrize("section", REQUIRED_SECTIONS)
def test_report_all_sections_present(section):
    assert section in build()


def test_limitations_always_present_even_when_clean():
    """The section must exist unconditionally. A report that omits what could
    not be checked invites the reader to assume full coverage."""
    d = build(reviews=[review("REV-1", "B1")],
              verifications=verif("nppes", VERIFIED) + verif("pecos", VERIFIED)
              + verif("oig_leie", VERIFIED) + verif("sam_gov", VERIFIED)
              + verif("rce_directory", VERIFIED) + verif("state_registry", VERIFIED)
              + verif("irs", VERIFIED))
    assert isinstance(d["limitations"], list)
    assert len(d["limitations"]) >= 1


def test_limitations_names_uncovered_sources():
    d = build()
    text = " ".join(d["limitations"]).lower()
    assert "sam_gov" in text
    assert "state_registry" in text


def test_limitations_lists_pending_b3_review_ids():
    d = build()
    text = " ".join(d["limitations"])
    assert "REV-2026-000004" in text


# ── arithmetic ───────────────────────────────────────────────────────────────

def test_b1b4_counts_correct():
    c = build()["classification_distribution"]["counts"]
    assert c == {"B1": 2, "B2": 1, "B3": 1, "B4": 1}


def test_discrepancy_rate_math():
    """Anything not B1 is a discrepancy — B3 included. 'Unexplained' is a
    finding, not a pass."""
    d = build()
    assert d["executive_summary"]["discrepancies_found"] == 3
    assert d["discrepancy_rate"]["rate"] == pytest.approx(3 / 5)


def test_confidence_interval_bounds_are_valid():
    ci = build()["discrepancy_rate"]
    assert 0.0 <= ci["lower"] <= ci["rate"] <= ci["upper"] <= 1.0
    assert ci["method"] == "wilson"


def test_reclassified_review_counts_in_its_resolved_bucket():
    """A human resolution is the finding of record, not the engine's first guess."""
    reviews = [review("REV-1", "B3", resolution="reclassified", reclass="B1"),
               review("REV-2", "B3")]
    c = build(reviews=reviews)["classification_distribution"]["counts"]
    assert c["B1"] == 1 and c["B3"] == 1


def test_unavailable_source_is_not_counted_as_a_finding():
    cov = build()["verification_coverage"]
    assert cov["pecos"][UNAVAILABLE] == 2
    assert cov["pecos"][NOT_FOUND] == 0
    note = build()["methodology"]["unavailable_handling"].lower()
    assert "does not count against" in note.replace("  ", " ")


def test_five_states_reported_separately():
    v = (verif("nppes", VERIFIED) + verif("nppes", NOT_FOUND)
         + verif("nppes", UNAVAILABLE) + verif("nppes", NOT_CHECKED)
         + verif("nppes", FAILED))
    cov = build(verifications=v)["verification_coverage"]["nppes"]
    assert all(cov[s] == 1 for s in (VERIFIED, NOT_FOUND, UNAVAILABLE,
                                     NOT_CHECKED, FAILED))


# ── edges ────────────────────────────────────────────────────────────────────

def test_empty_period_graceful():
    """Zeros must be explained, or they read as a clean bill of health."""
    d = build(reviews=[], verifications=[])
    assert d["executive_summary"]["entities_reviewed"] == 0
    assert d["discrepancy_rate"]["rate"] is None
    assert any("no reviews" in x.lower() for x in d["limitations"])


def test_configuration_records_rule_set_version():
    assert build()["configuration"]["rule_set_version"] == 1


def test_report_ids_are_stable_and_typed():
    assert report_id_for("weekly", date(2026, 8, 2)).startswith("WR-2026-W")
    assert report_id_for("quarterly", date(2026, 8, 2)) == "QR-2026-Q3"
    assert report_id_for("priority", date(2026, 8, 1)) == "PR-2026-08-01"


# ── rendering ────────────────────────────────────────────────────────────────

def test_html_contains_every_section_and_the_contract():
    html = render_html(build(), "WR-2026-W31")
    assert "7571MN26F80064" in html
    for heading in ("Executive Summary", "Classification Distribution",
                    "Limitations and Exceptions", "Configuration Used",
                    "Verification Coverage"):
        assert heading in html
    assert "REV-2026-000004" in html          # review ids surfaced to the reader


def test_html_escapes_injected_content():
    reviews = [review("<script>alert(1)</script>", "B1")]
    html = render_html(build(reviews=reviews), "WR-2026-W31")
    assert "<script>alert(1)</script>" not in html
    assert "&lt;script&gt;" in html


def test_bucket_labels_cover_all_four():
    assert set(BUCKET_LABELS) == {"B1", "B2", "B3", "B4"}


def test_build_limitations_never_returns_empty():
    assert build_limitations({}, [review("REV-1", "B1")]) != []


# ── P1.1 quarterly trend ─────────────────────────────────────────────────────

def test_weekly_trend_present_only_on_quarterly():
    from app.tefca_registry.report_generator import weekly_trend
    reviews = [dict(review("REV-1", "B1"), created_at="2026-07-01T10:00:00"),
               dict(review("REV-2", "B3"), created_at="2026-07-01T11:00:00"),
               dict(review("REV-3", "B1"), created_at="2026-07-08T10:00:00")]
    q = build_report_data(report_type="quarterly", period_start=date(2026, 7, 1),
                          period_end=date(2026, 9, 30), reviews=reviews,
                          verifications=[], rule_set_version=1)
    w = build_report_data(report_type="weekly", period_start=START,
                          period_end=END, reviews=reviews, verifications=[],
                          rule_set_version=1)
    assert "weekly_trend" in q
    assert "weekly_trend" not in w
    assert len(q["weekly_trend"]) == 2


def test_weekly_trend_counts_resolved_bucket():
    from app.tefca_registry.report_generator import weekly_trend
    t = weekly_trend([dict(review("REV-1", "B3", resolution="reclassified",
                                  reclass="B1"), created_at="2026-07-01T10:00:00")])
    assert t[0]["b1"] == 1 and t[0]["b3"] == 0


def test_weekly_trend_surfaces_undated_rather_than_dropping():
    """Silently dropping undated reviews would make the trend disagree with the
    distribution for no visible reason."""
    from app.tefca_registry.report_generator import weekly_trend
    t = weekly_trend([review("REV-1", "B1")])
    assert t and t[-1]["week"] == "undated" and t[-1]["b1"] == 1


# ── P1.2 Excel export ────────────────────────────────────────────────────────

def test_weekly_excel_has_three_sheets_and_opens():
    import io
    from openpyxl import load_workbook
    from app.tefca_registry.report_excel import build_weekly_excel
    data = build()
    blob = build_weekly_excel(data, "WR-2026-W31", [{
        "review_id": "REV-2026-000001", "entity_name": "Test Hospital",
        "npi": "1477978807", "entity_type": "participant",
        "verification": {"nppes": {"status": "verified"},
                         "pecos": {"status": "verified"},
                         "oig_leie": {"status": "clear"}},
        "bucket": "B1", "rule_code": "RULE-001", "rationale": "clean"}])
    assert blob[:2] == b"PK"
    wb = load_workbook(io.BytesIO(blob))
    assert "Entity Results" in wb.sheetnames
    assert "Summary Statistics" in wb.sheetnames
    assert "Limitations" in wb.sheetnames


def test_excel_limitations_sheet_is_never_empty():
    """If a reader opens the Excel and not the HTML they must still see the
    caveats, or the export launders a caveated report into a clean one."""
    import io
    from openpyxl import load_workbook
    from app.tefca_registry.report_excel import build_weekly_excel
    wb = load_workbook(io.BytesIO(build_weekly_excel(build(), "WR-2026-W31", [])))
    s = wb["Limitations"]
    assert s.max_row >= 5
    assert s.cell(5, 1).value


def test_excel_empty_period_says_so_rather_than_looking_broken():
    import io
    from openpyxl import load_workbook
    from app.tefca_registry.report_excel import build_weekly_excel
    wb = load_workbook(io.BytesIO(build_weekly_excel(
        build(reviews=[], verifications=[]), "WR-2026-W31", [])))
    ws = wb["Entity Results"]
    assert "No reviews" in str(ws.cell(2, 2).value)


def test_excel_reader_affordances():
    import io
    from openpyxl import load_workbook
    from app.tefca_registry.report_excel import build_weekly_excel
    wb = load_workbook(io.BytesIO(build_weekly_excel(build(), "WR-2026-W31", [])))
    ws = wb["Entity Results"]
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref.startswith("A1:")
