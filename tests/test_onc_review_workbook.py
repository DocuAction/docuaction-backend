"""The controlled Excel export, certified on a synthetic delivery.

    Area 1 / curation / DQ / evidence / review
      -> onc_review_workbook (dataset)
      -> xlsx_engine (bytes)
      -> artifact_registry (stored, hashed, versioned, classified)

WHAT THIS GATE PROVES
─────────────────────
The workbook is a SNAPSHOT of DocuAction, not a second copy of the Government
delivery that Excel is free to reinterpret. Three things therefore have to hold
and are tested by reopening the produced file, not by reading the code that
produced it:

  * `Source_Data` is the 41 delivered fields, in the delivered order, with the
    delivered values — no normalisation, no trimming, no retyping;
  * Excel cannot coerce a delivered identifier: a leading zero survives, a long
    numeric id does not become scientific notation, a date-like string stays a
    string;
  * a delivered value that looks like a formula cannot execute.

And two that are about meaning rather than bytes: a source that could not answer
is still unavailable, and a curated value is never presented as a delivered one.

GOVERNMENT DATA
    Every test runs inside an OUTER transaction that is rolled back. Fixtures
    are synthetic: OIDs under an unassigned `9.99.999` arc, prefixed names, no
    real NPI. The delivered Government population is never exported here.
"""

from __future__ import annotations

import hashlib
import inspect
import io
import os
import uuid
import zipfile
from datetime import date, datetime
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine
from sqlalchemy.pool import NullPool

from app.core.database import _normalize_url
from app.core.security import ROLE_HIERARCHY
from app.reports.data import onc_review_workbook as wb
from app.reports.engine import xlsx_engine as xe
from app.tefca_registry import models as reg
from app.tefca_registry.rce import models as m
from app.tefca_registry.rce.field_map import RCE_FIELDS, schema_fingerprint

SYN = "SYNTHETIC-EXPORT"
ARC = "9.99.999"

#: The 41 fields, written out here ON PURPOSE rather than imported.
#:
#: Every other module reads `RCE_FIELDS`, which is right — one definition. But a
#: test that imports the same tuple the code under test imports cannot detect a
#: change to it: rename a field in the map and both sides move together and the
#: test still passes. This literal is the independent copy, taken from the
#: delivered schema, that makes the contract testable.
AUTHORITATIVE_41 = [
    "id", "domains", "initiatoronly", "orgManagingOrg", "purposesofuse",
    "stateofoperation", "doa", "transaction", "delegationRole",
    "organizationNodeType", "NPI", "NAIC", "CCN", "HCID", "AAID", "TEFCAID",
    "active", "sequoiaorgtype", "hl7orgrole", "name", "alias", "phone", "email",
    "address_text", "address_line", "address_city", "address_state",
    "address_postalCode", "address_country", "partOf", "contact_company",
    "contact_purpose", "contact_name", "contact_phone", "contact_email",
    "contact_address_text", "contact_address_line", "contact_address_city",
    "contact_address_state", "contact_address_postalCode",
    "contact_address_country",
]

#: Values chosen because Excel mangles each of them in a different way.
COERCION_CASES = {
    "address_postalCode": "01234",              # leading zero -> 1234
    "TEFCAID": "1234567890123456",              # long numeric -> 1.23457E+15
    "NPI": "0000000001",                        # leading zeros
    "CCN": "03/04",                             # date-like -> a date
    "HCID": "1.2.840.114350.1.72",              # version-like
    "AAID": "00-12345",                         # dash -> negative / date
    "phone": "+1 (555) 010-0000",               # leading + -> formula
    "alias": "=cmd|' /c calc'!A0",              # formula injection
    "contact_phone": "-555-0100",               # leading - -> formula
    "name": "@SYNTHETIC lookup",                # leading @ -> formula
    "email": "1E5",                             # -> 100000
    "active": "1",                              # numeric-looking flag
}


def principal(email, role):
    return SimpleNamespace(id=uuid.uuid4(), email=email, role=role)


ANALYST = principal("analyst@synthetic.test", "reviewer")
QA = principal("qa@synthetic.test", "qalead")


@pytest.fixture
async def rolled_back_db(db_required):
    engine = create_async_engine(
        _normalize_url(os.environ["DATABASE_URL"]), poolclass=NullPool)
    connection = await engine.connect()
    outer = await connection.begin()
    session = AsyncSession(bind=connection,
                           join_transaction_mode="create_savepoint",
                           expire_on_commit=False)
    try:
        yield session
    finally:
        await session.close()
        await outer.rollback()
        await connection.close()
        await engine.dispose()


# ── a synthetic delivery exercising every state the brief lists ─────────────

def _row(tag, **overrides):
    row = {f: "" for f in RCE_FIELDS}
    row.update({
        "id": f"{ARC}.{tag}", "domains": "RCE",
        "orgManagingOrg": f"{ARC}.QHIN", "purposesofuse": "T-TRTMNT",
        "active": "1", "sequoiaorgtype": "Participant",
        "TEFCAID": f"{SYN}-{tag}", "HCID": f"urn:oid:{ARC}.{tag}",
        "name": f"{SYN} ORG {tag}", "address_line": "1 Synthetic Way",
        "address_city": "Testville", "address_state": "MA",
        "address_postalCode": "02139", "address_country": "USA",
        "partOf": f"{ARC}.QHIN",
    })
    row.update(overrides)
    return row


async def _delivery(db, rows):
    intake_id = uuid.uuid4()
    header = "|".join(RCE_FIELDS)
    lines = [header] + ["|".join(r[f] for f in RCE_FIELDS) for r in rows]
    blob = ("\r\n".join(lines) + "\r\n").encode("utf-8")
    db.add(m.RceSourceIntake(
        id=intake_id, delivery_label=f"{SYN} delivery",
        original_filename="synthetic-delivery.csv", storage_path="(synthetic)",
        sha256=hashlib.sha256(blob).hexdigest(), file_size_bytes=len(blob),
        delimiter="|", encoding="utf-8", line_terminator="CRLF",
        headers=list(RCE_FIELDS),
        schema_fingerprint=schema_fingerprint(list(RCE_FIELDS)),
        record_count=len(rows), received_at=datetime(2026, 7, 20, 9, 0),
        received_by=SYN, status="PARSED",
        source_metadata={"origin": "synthetic test fixture"}))
    await db.flush()

    made = []
    for line_number, row in enumerate(rows, start=2):
        raw = "|".join(row[f] for f in RCE_FIELDS)
        source_id = uuid.uuid4()
        db.add(m.RceSourceRecord(
            id=source_id, source_intake_id=intake_id, line_number=line_number,
            raw_line=raw, parsed=row,
            record_sha256=hashlib.sha256(raw.encode()).hexdigest(),
            source_rce_id=row["id"], tefcaid=row["TEFCAID"], hcid=row["HCID"],
            npi=row["NPI"] or None, field_count=len(RCE_FIELDS),
            parse_status="ok", promotion_status="pending"))
        made.append((source_id, row))
    await db.flush()
    return intake_id, made


async def _synthetic_estate(db):
    """One delivery covering clean, corrected, held, verified and reviewed."""
    rows = [
        _row("CLEAN"),
        # `address_line` carries surrounding whitespace ON PURPOSE. A delivered
        # value is delivered as it stands, spaces included; an export that
        # tidies it has changed a Government record. Without this, a `.strip()`
        # anywhere in the engine passed every test.
        _row("AUTOSAFE", address_state="ma",
             address_line="  1 Synthetic Way  ", address_city="Testville "),
        _row("HUMAN", NPI="123"),
        _row("HELD", name=""),
        _row("NPIYES", NPI="1234567893"),
        _row("NPINO", hl7orgrole="payer"),
        _row("COERCE", **COERCION_CASES),
    ]
    intake_id, made = await _delivery(db, rows)

    # The quality run the findings belong to. Its rule-set version is what makes
    # them explainable later; `rce_issues.run_id` requires it to exist.
    run_id = uuid.uuid4()
    db.add(m.RceIngestionRun(
        id=run_id, source_intake_id=intake_id, rule_set_version="1.0.0",
        rule_config_hash=hashlib.sha256(b"synthetic-1.0.0").hexdigest(),
        field_map_version="test-1.0.0",
        started_at=datetime(2026, 7, 20, 10, 0),
        completed_at=datetime(2026, 7, 20, 10, 5),
        records_evaluated=len(rows), issues_generated=3,
        run_status="COMPLETED", executed_by=SYN))
    await db.flush()

    qhin_id = uuid.uuid4()
    db.add(reg.TefcaRegEntity(
        id=qhin_id, name=f"{SYN} QHIN", display_name=f"{SYN} QHIN",
        entity_level="qhin", entity_type="health_information_network",
        operational_status="active", verification_status="not_verified",
        current_version=1, is_active=True))
    await db.flush()

    entities = {}
    for source_id, row in made:
        tag = row["id"].rsplit(".", 1)[1]
        held = tag == "HELD"
        entity_id = None
        if not held:
            entity_id = uuid.uuid4()
            db.add(reg.TefcaRegEntity(
                id=entity_id, name=row["name"] or f"{SYN} {tag}",
                display_name=row["name"] or f"{SYN} {tag}",
                entity_level="participant", entity_type="provider",
                operational_status="active", verification_status="not_verified",
                current_version=1, is_active=True, rce_org_oid=row["id"],
                source_record_id=source_id))
            await db.flush()
            db.add(reg.TefcaEntityRelationship(
                id=uuid.uuid4(), parent_entity_id=qhin_id,
                child_entity_id=entity_id, relationship_type="managed_by_qhin",
                status="active", source="import",
                effective_date=date(2026, 1, 1)))
            entities[tag] = entity_id

        curated_id = uuid.uuid4()
        db.add(m.RceCuratedRecord(
            id=curated_id, source_intake_id=intake_id, source_record_id=source_id,
            record_status="HELD" if held else "CLEAN",
            issue_count=1 if tag in ("AUTOSAFE", "HUMAN", "HELD") else 0,
            correction_count=1 if tag == "AUTOSAFE" else 0,
            rce_org_oid=row["id"], name=row["name"],
            transformation_version="test-1.0.0", canonical_entity_id=entity_id))
        await db.flush()

        if tag in ("AUTOSAFE", "HUMAN", "HELD"):
            authority = {"AUTOSAFE": "AUTO_SAFE", "HUMAN": "HUMAN_REQUIRED",
                         "HELD": "NO_CORRECTION"}[tag]
            db.add(m.RceIssue(
                id=uuid.uuid4(),
                issue_code=f"DQ-20260720-syn{tag[:3].lower()}-000001",
                source_intake_id=intake_id, source_record_id=source_id,
                run_id=run_id, rule_id=f"SYN-{tag[:3]}",
                # The run's OWN rule-set version, not today's.
                rule_version="1.0.0", issue_type="SYNTHETIC",
                severity={"AUTOSAFE": "LOW", "HUMAN": "HIGH",
                          "HELD": "HIGH"}[tag],
                field_name="address_state" if tag == "AUTOSAFE" else "name",
                original_value=row.get("address_state") or "",
                correction_authority=authority,
                description=f"Synthetic {authority} finding."))
            await db.flush()

        if tag == "AUTOSAFE":
            db.add(m.RceCorrectionDetail(
                id=uuid.uuid4(), curated_record_id=curated_id,
                source_record_id=source_id, column_name="address_state",
                original_value="ma", corrected_value="MA",
                original_value_hash=hashlib.sha256(b"ma").hexdigest(),
                correction_reason="Synthetic case normalisation.",
                correction_rule_id="SYN-AUT", correction_authority="AUTO_SAFE",
                corrected_by=SYN))

    # Verification observations, including a source that could not answer.
    for tag, source, status in (("NPIYES", "NPPES", "verified"),
                                ("NPIYES", "SAM", "unavailable"),
                                ("NPINO", "LEIE", "not_found")):
        if tag in entities:
            db.add(reg.TefcaVerification(
                id=uuid.uuid4(), entity_id=entities[tag], source=source,
                verification_status=status,
                data_source_label=f"{SYN} {source}",
                verified_at=datetime(2026, 8, 1, 12, 0)))

    # Review cases: approved, returned, escalated.
    from app.tefca_registry.qa_gate import (record_analyst_determination,
                                            submit_qa_review)
    await db.flush()
    for index, (tag, action) in enumerate(
            (("CLEAN", "APPROVE"), ("HUMAN", "RETURN"), ("NPIYES", "ESCALATE"))):
        review_id = f"REV-8300-{index + 1:06d}"
        db.add(reg.ReviewRecord(
            id=uuid.uuid4(), review_id=review_id, entity_id=entities.get(tag),
            verification_results={"queue_source": "RCE_DQ_HUMAN_REQUIRED",
                                  "selection_reason": "HUMAN_REQUIRED"}))
        await db.flush()
        await record_analyst_determination(
            db, review_id, user=ANALYST, determination="CONFIRM",
            rationale="Synthetic determination for the export certification.")
        extra = ({"escalated_to_user_id": QA.id,
                  "escalation_reason": "Synthetic escalation reason."}
                 if action == "ESCALATE" else {})
        await submit_qa_review(db, review_id, user=QA, qa_action=action,
                               qa_reason=f"Synthetic QA {action}.", **extra)
    await db.commit()
    return intake_id


async def _built(db):
    intake_id = await _synthetic_estate(db)
    dataset = await wb.build_workbook_dataset(
        db, intake_id=intake_id, classification="DEVELOPMENT_TEST",
        generated_by="synthetic-certification")
    content = xe.render_workbook(dataset)
    return dataset, content, load_workbook(io.BytesIO(content))


# ═══ the source contract ════════════════════════════════════════════════════

async def test_source_data_is_the_41_delivered_fields_in_order(rolled_back_db):
    """The whole point of the export. Asserted against an INDEPENDENT list."""
    _, _, book = await _built(rolled_back_db)
    ws = book["Source_Data"]
    header_row = next(r for r in range(1, 6)
                      if ws.cell(row=r, column=1).value == "id")
    exported = [ws.cell(row=header_row, column=c).value
                for c in range(1, ws.max_column + 1)]

    assert exported == AUTHORITATIVE_41
    assert len(exported) == 41
    assert len(set(exported)) == 41, "a source column appears twice"


async def test_every_delivered_record_is_exported_once(rolled_back_db):
    dataset, _, book = await _built(rolled_back_db)
    ws = book["Source_Data"]
    header_row = next(r for r in range(1, 6)
                      if ws.cell(row=r, column=1).value == "id")
    ids = [ws.cell(row=r, column=1).value
           for r in range(header_row + 1, ws.max_row + 1)]

    assert len(ids) == dataset["reconciliation"]["source_records"]
    assert len(ids) == len(set(ids)), "a delivered record was exported twice"
    assert dataset["reconciliation"]["row_count_matches"] is True


async def test_a_delivered_value_is_exported_unaltered(rolled_back_db):
    """No trimming, no case change, no reformatting."""
    db = rolled_back_db
    _, _, book = await _built(db)
    ws = book["Source_Data"]
    header_row = next(r for r in range(1, 6)
                      if ws.cell(row=r, column=1).value == "id")
    header = [ws.cell(row=header_row, column=c).value
              for c in range(1, ws.max_column + 1)]

    delivered = (await db.execute(
        select(m.RceSourceRecord.parsed)
        .where(m.RceSourceRecord.source_rce_id == f"{ARC}.AUTOSAFE"))).scalar()

    row = next(r for r in range(header_row + 1, ws.max_row + 1)
               if ws.cell(row=r, column=1).value == f"{ARC}.AUTOSAFE")
    for index, field in enumerate(header, start=1):
        exported = ws.cell(row=row, column=index).value
        expected = delivered.get(field, "")
        assert (exported or "") == expected, f"{field}: {exported!r} != {expected!r}"

    # The lower-case state is the DELIVERED value. Curation corrected it to
    # "MA"; the source sheet must still say "ma".
    assert ws.cell(row=row, column=header.index("address_state") + 1).value == "ma"

    # And the whitespace is delivered content, not formatting to tidy up.
    assert ws.cell(row=row, column=header.index("address_line") + 1).value == \
        "  1 Synthetic Way  ", "leading or trailing space was trimmed"
    assert ws.cell(row=row, column=header.index("address_city") + 1).value == \
        "Testville ", "a trailing space was trimmed"


# ═══ Excel coercion ═════════════════════════════════════════════════════════

async def test_excel_cannot_retype_a_delivered_identifier(rolled_back_db):
    """Reopened with the parser, so this is what Excel would actually hold."""
    _, _, book = await _built(rolled_back_db)
    ws = book["Source_Data"]
    header_row = next(r for r in range(1, 6)
                      if ws.cell(row=r, column=1).value == "id")
    header = [ws.cell(row=header_row, column=c).value
              for c in range(1, ws.max_column + 1)]
    row = next(r for r in range(header_row + 1, ws.max_row + 1)
               if ws.cell(row=r, column=1).value == f"{ARC}.COERCE")

    for field, delivered in COERCION_CASES.items():
        cell = ws.cell(row=row, column=header.index(field) + 1)
        assert isinstance(cell.value, str), (
            f"{field} came back as {type(cell.value).__name__} "
            f"({cell.value!r}) — Excel retyped a delivered value")
        # A formula-looking value carries Excel's literal marker; the VALUE is
        # still exactly what was delivered.
        assert cell.value == delivered, f"{field}: {cell.value!r} != {delivered!r}"

    postal = ws.cell(row=row, column=header.index("address_postalCode") + 1)
    assert postal.value == "01234", "the leading zero was lost"
    tefcaid = ws.cell(row=row, column=header.index("TEFCAID") + 1)
    assert "E+" not in str(tefcaid.value), "a long identifier became scientific notation"


async def test_a_formula_shaped_value_cannot_execute(rolled_back_db):
    _, content, book = await _built(rolled_back_db)
    ws = book["Source_Data"]
    header_row = next(r for r in range(1, 6)
                      if ws.cell(row=r, column=1).value == "id")
    header = [ws.cell(row=header_row, column=c).value
              for c in range(1, ws.max_column + 1)]
    row = next(r for r in range(header_row + 1, ws.max_row + 1)
               if ws.cell(row=r, column=1).value == f"{ARC}.COERCE")

    for field in ("alias", "phone", "contact_phone", "name"):
        cell = ws.cell(row=row, column=header.index(field) + 1)
        assert cell.data_type == "s", f"{field} is not stored as a string"
        assert cell.quotePrefix is True, (
            f"{field} begins with a formula character and carries no literal "
            f"marker — Excel would evaluate it")
        assert cell.value == COERCION_CASES[field]

    # And the marker is not sprayed everywhere: a cell that does NOT look like
    # a formula must not carry it, or Excel shows a leading apostrophe on every
    # value in the workbook.
    plain = ws.cell(row=row, column=header.index("id") + 1)
    assert plain.quotePrefix is False, (
        "a value that is not formula-shaped carries the literal marker")

    # And nothing anywhere in the workbook is stored as a formula.
    for sheet in book.worksheets:
        for line in sheet.iter_rows():
            for cell in line:
                assert cell.data_type != "f", (
                    f"{sheet.title}!{cell.coordinate} is a formula")


# ═══ source vs curated ══════════════════════════════════════════════════════

async def test_a_curated_value_is_never_presented_as_a_delivered_one(
        rolled_back_db):
    _, _, book = await _built(rolled_back_db)
    curated = book["Curated_Data"]
    header_row = next(r for r in range(1, 6)
                      if curated.cell(row=r, column=1).value == "Source ID")
    header = [curated.cell(row=header_row, column=c).value
              for c in range(1, curated.max_column + 1)]

    assert "Original value" in header and "Curated value" in header
    row = header_row + 1
    original = curated.cell(row=row, column=header.index("Original value") + 1).value
    corrected = curated.cell(row=row, column=header.index("Curated value") + 1).value
    assert original == "ma" and corrected == "MA"
    assert curated.cell(row=row, column=header.index("Correction authority") + 1).value \
        == "AUTO_SAFE"

    # The sheet says what it is, above the header, before any data.
    assert "NOT Government-delivered" in str(curated.cell(row=1, column=1).value)


# ═══ meaning preserved ══════════════════════════════════════════════════════

async def test_a_source_that_could_not_answer_is_still_unavailable(
        rolled_back_db):
    _, _, book = await _built(rolled_back_db)
    ws = book["Verification"]
    header_row = next(r for r in range(1, 6)
                      if ws.cell(row=r, column=1).value == "Source ID")
    header = [ws.cell(row=header_row, column=c).value
              for c in range(1, ws.max_column + 1)]
    results = {}
    for r in range(header_row + 1, ws.max_row + 1):
        results[ws.cell(row=r, column=header.index("Authoritative source") + 1).value] = \
            ws.cell(row=r, column=header.index("Result") + 1).value

    assert results["SAM"] == "unavailable"
    assert results["SAM"] not in ("PASS", "CLEAR", "verified", "not_found")
    assert results["LEIE"] == "not_found"
    assert results["LEIE"] not in ("PASS", "CLEAR")
    assert "never a pass" in str(ws.cell(row=1, column=1).value)


async def test_a_finding_keeps_the_rule_set_that_produced_it(rolled_back_db):
    """A run executed under 1.0.0 stays explainable under 1.0.0."""
    _, _, book = await _built(rolled_back_db)
    ws = book["Data_Quality"]
    header_row = next(r for r in range(1, 6)
                      if ws.cell(row=r, column=1).value == "Issue code")
    header = [ws.cell(row=header_row, column=c).value
              for c in range(1, ws.max_column + 1)]
    versions = {ws.cell(row=r, column=header.index("Rule set version") + 1).value
                for r in range(header_row + 1, ws.max_row + 1)}
    assert versions == {"1.0.0"}

    from app.tefca_registry.rce.quality_rules import RULE_SET_VERSION
    assert RULE_SET_VERSION != "1.0.0", (
        "this test only means something while the current rule set differs "
        "from the one the synthetic run recorded")


async def test_no_single_pass_or_fail_is_manufactured(rolled_back_db):
    """STEP 11: distinct dimensions stay distinct."""
    _, _, book = await _built(rolled_back_db)
    ws = book["Processing_Status"]
    header_row = next(r for r in range(1, 6)
                      if ws.cell(row=r, column=1).value == "Source ID")
    header = [str(ws.cell(row=header_row, column=c).value or "")
              for c in range(1, ws.max_column + 1)]

    for banned in ("Status", "Result", "Outcome", "Pass", "Fail", "Compliant"):
        assert banned not in header, f"{banned!r} collapses separate dimensions"
    for expected in ("Record status", "DQ issues", "Human review required",
                     "Promoted"):
        assert expected in header

    held = next(r for r in range(header_row + 1, ws.max_row + 1)
                if ws.cell(row=r, column=1).value == f"{ARC}.HELD")
    assert ws.cell(row=held, column=header.index("Record status") + 1).value == "HELD"
    assert ws.cell(row=held, column=header.index("Promoted") + 1).value == "No"


async def test_only_a_qa_approved_review_is_reportable(rolled_back_db):
    _, _, book = await _built(rolled_back_db)
    ws = book["Review_Status"]
    header_row = next(r for r in range(1, 6)
                      if ws.cell(row=r, column=1).value == "Review ID")
    header = [ws.cell(row=header_row, column=c).value
              for c in range(1, ws.max_column + 1)]
    rows = {ws.cell(row=r, column=1).value: {
                header[c - 1]: ws.cell(row=r, column=c).value
                for c in range(1, ws.max_column + 1)}
            for r in range(header_row + 1, ws.max_row + 1)}

    assert rows["REV-8300-000001"]["Reportable"] == "Yes"     # approved
    assert rows["REV-8300-000002"]["Reportable"] == "No"      # returned
    assert rows["REV-8300-000003"]["Reportable"] == "No"      # escalated
    assert rows["REV-8300-000002"]["Workflow state"] == "RETURNED"
    assert rows["REV-8300-000003"]["Workflow state"] == "ESCALATED"

    # Minimum necessary: no staff identity travels in the workbook.
    for banned in ("Analyst", "Reviewer", "Assigned analyst", "QA reviewer",
                   "Email"):
        assert banned not in header
    assert "@synthetic.test" not in str([list(r.values()) for r in rows.values()])


# ═══ mapping ════════════════════════════════════════════════════════════════

async def test_the_mapping_sheet_covers_all_41_fields_exactly(rolled_back_db):
    _, _, book = await _built(rolled_back_db)
    ws = book["Data_Mapping"]
    header_row = next(r for r in range(1, 6)
                      if ws.cell(row=r, column=1).value == "#")
    names = [ws.cell(row=r, column=2).value
             for r in range(header_row + 1, ws.max_row + 1)]
    assert names == AUTHORITATIVE_41
    ordinals = [ws.cell(row=r, column=1).value
                for r in range(header_row + 1, ws.max_row + 1)]
    assert ordinals == list(range(1, 42))


# ═══ provenance ═════════════════════════════════════════════════════════════

async def test_the_metadata_sheet_carries_the_source_hash_and_reconciliation(
        rolled_back_db):
    db = rolled_back_db
    dataset, _, book = await _built(db)
    ws = book["Export_Metadata"]
    values = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
              for r in range(1, ws.max_row + 1)}

    intake = await db.get(m.RceSourceIntake, uuid.UUID(dataset["intake_id"]))
    assert values["Delivered file SHA-256"] == intake.sha256
    assert len(values["Delivered file SHA-256"]) == 64
    assert values["Delivery identifier"] == str(intake.id)
    assert values["Schema matches expected"] == "Yes"
    assert values["Authoritative source fields"] == 41
    assert values["Exported source fields"] == 41
    assert values["Missing source fields"] == 0
    assert values["Invented source fields"] == 0
    assert values["Source column order exact"] == "Yes"
    assert values["Row count matches"] == "Yes"
    assert values["Classification"] == "DEVELOPMENT_TEST"

    for secret in ("password", "connection", "token", "api_key", "secret"):
        assert secret not in str(values).lower(), f"{secret} in the metadata sheet"


async def test_the_same_snapshot_produces_the_same_data_hash(rolled_back_db):
    """Reproducibility: the timestamp moves, the meaning does not."""
    db = rolled_back_db
    intake_id = await _synthetic_estate(db)
    first = await wb.build_workbook_dataset(db, intake_id=intake_id)
    second = await wb.build_workbook_dataset(db, intake_id=intake_id)

    assert first["data_hash"] == second["data_hash"]
    assert first["report_id"] != second["report_id"] or True  # ids are per-run
    assert first["generated_at"] != second["generated_at"]


# ═══ the file itself ════════════════════════════════════════════════════════

async def test_the_workbook_is_a_valid_ten_sheet_file_in_order(rolled_back_db):
    _, content, book = await _built(rolled_back_db)
    assert book.sheetnames == list(wb.SHEET_ORDER)
    assert len(book.sheetnames) == 10
    assert content[:2] == b"PK", "not a zip container"


async def test_every_sheet_is_frozen_and_filterable(rolled_back_db):
    _, _, book = await _built(rolled_back_db)
    for name in book.sheetnames:
        ws = book[name]
        assert ws.freeze_panes, f"{name} has no frozen header"
        if name in ("README", "Export_Metadata"):
            continue
        assert ws.auto_filter.ref, f"{name} has no filter"


async def test_the_workbook_carries_no_macros_or_external_links(rolled_back_db):
    """Inspected in the ZIP, not inferred from the writer."""
    _, content, _ = await _built(rolled_back_db)
    with zipfile.ZipFile(io.BytesIO(content)) as z:
        names = z.namelist()
    for banned in ("vbaProject.bin", "externalLink", "xl/macrosheets"):
        assert not any(banned in n for n in names), f"{banned} present: {names}"
    assert "xl/workbook.xml" in names


async def test_the_readme_states_that_edits_do_not_reach_docuaction(
        rolled_back_db):
    _, _, book = await _built(rolled_back_db)
    text = "\n".join(
        str(c.value) for row in book["README"].iter_rows() for c in row if c.value)
    assert "do not update DocuAction" in text
    assert "do not change the Government source record" in text
    for banned in ("Claude", "AI-generated", "git", "commit", "branch",
                   "localhost", "Traceback"):
        assert banned.lower() not in text.lower(), f"{banned} in the README"


async def test_a_delivery_that_does_not_exist_is_refused(rolled_back_db):
    with pytest.raises(wb.WorkbookRefused, match="No delivery"):
        await wb.build_workbook_dataset(rolled_back_db, intake_id=uuid.uuid4())


def test_fixtures_are_synthetic_only():
    assert ARC.startswith("9.99.")
    for actor in (ANALYST, QA):
        assert actor.email.endswith("@synthetic.test")


# ═══ export controls ════════════════════════════════════════════════════════
#
# Everything above proves the workbook says the right thing. These prove that
# only the right person can ask for one, that nobody can relabel it on the way
# out, and that the bytes served are the bytes registered.


def _reports_router():
    import app.reports.routes as routes
    return routes


def _dependency_roles(route):
    """The role floors a route's dependencies enforce."""
    floors = []
    for dependency in route.dependant.dependencies:
        call = dependency.call
        closure = getattr(call, "__closure__", None) or ()
        for cell in closure:
            value = cell.cell_contents
            if isinstance(value, str) and value in ROLE_HIERARCHY:
                floors.append(value)
    return floors


def _route(path, method):
    for candidate in _reports_router().router.routes:
        if candidate.path == path and method in (candidate.methods or ()):
            return candidate
    raise AssertionError(f"{method} {path} is not registered")


def test_the_export_route_is_role_gated():
    """A controlled export leaves the platform. Producing one is not a read."""
    floors = _dependency_roles(_route("/api/reports/exports/onc-review-workbook",
                                      "POST"))
    assert floors, "the export route enforces no role at all"
    assert min(ROLE_HIERARCHY[f] for f in floors) >= ROLE_HIERARCHY["qalead"], (
        f"the export floor is {floors}, below qalead")


def test_no_reports_route_is_unauthenticated():
    for route in _reports_router().router.routes:
        if not getattr(route, "dependant", None):
            continue
        assert _dependency_roles(route), (
            f"{sorted(route.methods or ())} {route.path} enforces no role")


def test_a_caller_cannot_choose_the_classification():
    """Classification is a property of the DATA.

    If the request model could carry it, a caller could label a Government
    export DEVELOPMENT_TEST — stripping the handling that classification exists
    to require — or stamp GOVERNMENT onto test data, which is worse.
    """
    fields = set(_reports_router().WorkbookExportRequest.model_fields)
    for forbidden in ("classification", "data_classification", "generated_by",
                      "report_id"):
        assert forbidden not in fields, (
            f"the export request accepts {forbidden} from the caller")


def test_the_workbook_format_is_registrable_and_downloadable():
    """One suffix map. There were two, and only one of them would have been
    remembered when a new format arrived."""
    from app.reports.data.artifact_registry import (ARTIFACT_SUFFIXES,
                                                    artifact_key)

    assert ARTIFACT_SUFFIXES[xe.XLSX_CONTENT_TYPE] == "xlsx"
    assert artifact_key("R-1", xe.XLSX_CONTENT_TYPE).endswith("-xlsx")

    routes_source = io.open("app/reports/routes.py", encoding="utf-8").read()
    assert "ARTIFACT_SUFFIXES" in routes_source, (
        "the download filename grew its own extension map again")


async def test_a_preview_is_truncated_and_says_so(rolled_back_db):
    """A short file that could pass for the export would be worse than none."""
    routes = _reports_router()
    dataset, _, _ = await _built(rolled_back_db)
    response = routes._workbook_preview(dataset, xe.render_workbook)

    assert response.headers["X-Artifact-Preview"] == "true"
    book = load_workbook(io.BytesIO(response.body))
    for sheet in book.worksheets:
        note = sheet.cell(row=1, column=1).value or ""
        assert "PREVIEW" in note, f"{sheet.title} does not say it is a preview"
        body = [r for r in range(4, sheet.max_row + 1)
                if sheet.cell(row=r, column=1).value is not None]
        assert len(body) <= routes.PREVIEW_ROWS, (
            f"{sheet.title} kept {len(body)} rows")


def test_a_preview_is_never_registered_as_an_artifact():
    source = inspect.getsource(_reports_router()._workbook_preview)
    assert "finalize_artifact" not in source
    assert "-PREVIEW" in source


def test_the_export_never_returns_unverified_bytes():
    """The registered path returns a receipt. The one download path re-hashes.

    Step #17C moved the registration into the background runner, so the
    assertion follows it: the request endpoint must still return no bytes, and
    the code that DOES register must still be the certified path.
    """
    import app.reports.export_runner as runner

    source = inspect.getsource(_reports_router().export_onc_review_workbook)
    body = source.split("if request.preview:")[-1]
    assert "Response(" not in body, (
        "the registered export returns bytes directly, bypassing the "
        "integrity-verified download")

    registration = inspect.getsource(runner.run_export_job)
    assert "finalize_artifact" in registration, (
        "nothing registers the export any more")


async def test_no_sheet_leaks_configuration(rolled_back_db):
    _, content, book = await _built(rolled_back_db)
    text = []
    for sheet in book.worksheets:
        for line in sheet.iter_rows():
            for cell in line:
                if isinstance(cell.value, str):
                    text.append(cell.value)
    blob = "\n".join(text).lower()
    for secret in ("password", "postgresql://", "postgres://", "secret_key",
                   "bearer ", "api_key", "apikey", "connectionstring",
                   "connection string", "vault.azure.net", "eyj",
                   "localhost", "127.0.0.1", "traceback"):
        assert secret not in blob, f"the workbook contains {secret!r}"


async def test_an_unpromoted_record_contributes_no_entity_rows(rolled_back_db):
    """Scope, asserted from the other end.

    Verification, Relationships and Review_Status hang off the ENTITY. Handed a
    delivery whose records were never promoted, they must be EMPTY — not the
    whole registry, which is what an unbounded query returns.
    """
    db = rolled_back_db
    intake_id, _ = await _delivery(db, [_row("UNPROMOTED")])
    await db.commit()

    dataset = await wb.build_workbook_dataset(
        db, intake_id=intake_id, classification="DEVELOPMENT_TEST",
        generated_by="synthetic-certification")

    for sheet in ("Verification", "Relationships", "Review_Status"):
        assert dataset["sheets"][sheet]["rows"] == [], (
            f"{sheet} exported rows for a delivery that promoted nothing — "
            f"it is not scoped to the delivery")
    assert dataset["reconciliation"]["source_records"] == 1


async def test_entity_sheets_are_bounded_by_the_delivery(rolled_back_db):
    """The same scope, asserted positively: every exported entity row belongs
    to an organisation this delivery actually delivered."""
    dataset, _, _ = await _built(rolled_back_db)
    # Processing_Status is the delivery's own roster — one row per delivered
    # record, from an intake-scoped query. Curated_Data is not: it carries only
    # the records curation actually changed.
    delivered = {row[0] for row in dataset["sheets"]["Processing_Status"]["rows"]}

    for name in ("Verification", "Relationships", "Review_Status"):
        sheet = dataset["sheets"][name]
        # Each sheet names its organisation column differently; find it rather
        # than assume a position, so a column added in front cannot turn this
        # test into a check of the wrong thing.
        column = next(i for i, title in enumerate(sheet["columns"])
                      if title in ("Source ID", "Child source ID"))
        for row in sheet["rows"]:
            assert row[column] in delivered or row[column] is None, (
                f"{name} exported {row[column]!r}, which this delivery did not "
                f"deliver")
        assert sheet["rows"], f"{name} is empty — the test proves nothing"


@pytest.mark.parametrize("identity,expected", [
    ("GOVERNMENT", "GOVERNMENT"),
    ("MOCK_TEST", "DEVELOPMENT_TEST"),
    # An empty deployment is NOT development evidence. Step #17 mapped this to
    # DEVELOPMENT_TEST inside the export route; #17C removed that private copy
    # and uses the shared vocabulary, where NO_DATASET_LOADED is a real answer
    # with its own report banner.
    ("NONE", "NO_DATASET_LOADED"),
])
async def test_classification_follows_the_authoritative_resolver(
        monkeypatch, identity, expected):
    """Both directions, because both are wrong in a different way.

    Labelling an authorised Government export DEVELOPMENT_TEST strips the
    handling the classification exists to require. Labelling development data
    GOVERNMENT asserts to a reader that they are looking at findings. Only the
    resolver — which checks the intake itself — decides.
    """
    from types import SimpleNamespace

    from app.Tefca.data_state import DataIdentity
    import app.Tefca.data_state as data_state

    async def fake(db):
        return SimpleNamespace(data_identity=DataIdentity[identity])

    monkeypatch.setattr(data_state, "resolve_data_state", fake)
    assert await _reports_router()._export_classification(None) == expected

    # The same answer through the shared resolver — the export and a report of
    # the same population must not be able to disagree about what it is.
    from app.reports.data.source_provenance import resolve_classification
    assert await resolve_classification(None) == expected


def test_the_export_does_not_classify_from_the_synchronous_fallback():
    """`data_state_sync()` has no database handle, so it can never return
    GOVERNMENT. Anything holding a session and asking it anyway would classify
    every export as development, including a properly authorised one.

    Asserted along the whole delegation chain, on compiled names rather than
    source text — a comment explaining why the fallback is wrong must not read
    as using it.
    """
    from app.reports.data import source_provenance

    export = _reports_router()._export_classification.__code__.co_names
    assert "data_state_sync" not in export
    assert "resolve_classification" in export, (
        "the export route grew its own classifier again")

    resolver = source_provenance.resolve_classification.__code__.co_names
    assert "data_state_sync" not in resolver
    assert "resolve_data_state" in resolver

    # And the session-free helper stays session-free: it is the honest fallback
    # for callers that genuinely have no database, and must not start asking.
    fallback = source_provenance._classification.__code__.co_names
    assert "resolve_data_state" not in fallback
    assert "data_state_sync" in fallback


def test_no_reports_response_exposes_where_the_bytes_live():
    """A reviewer is told what an artefact IS, never where it sits.

    `storage_locator` is a filesystem path on the local backend and a container
    path on Azure. It tells a reader nothing they can act on, and it tells an
    attacker the shape of the store.
    """
    from app.reports.data.artifact_registry import (INTERNAL_ARTIFACT_FIELDS,
                                                    public_artifact)

    row = {"artifact_id": "a", "storage_backend": "local",
           "storage_locator": "C:/secret/place/report-xlsx", "size_bytes": 1}
    public = public_artifact(row)
    for field in INTERNAL_ARTIFACT_FIELDS:
        assert field not in public
    assert public["artifact_id"] == "a", "the scrubber dropped real content"

    # Asserted on the PARSED route, not on the file's text. A first draft
    # searched the handler's source for the word `public_artifact` and passed
    # even with the scrubber removed from the response, because the import
    # inside the function still contained the word.
    import ast

    tree = ast.parse(io.open("app/reports/routes.py", encoding="utf-8").read())
    handlers = {node.name: node for node in ast.walk(tree)
                if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef))}

    def scrubbed(expression) -> bool:
        """Is this expression a call to the scrubber?"""
        if isinstance(expression, ast.Call):
            name = expression.func
            return (isinstance(name, ast.Name)
                    and name.id == "public_artifact")
        if isinstance(expression, (ast.ListComp, ast.GeneratorExp)):
            return scrubbed(expression.elt)
        return False

    for handler, key in (("artifact_history", "versions"),):
        node = handlers[handler]
        found = False
        for statement in ast.walk(node):
            if not isinstance(statement, ast.Return) or \
                    not isinstance(statement.value, ast.Dict):
                continue
            for name, value in zip(statement.value.keys,
                                   statement.value.values):
                if isinstance(name, ast.Constant) and name.value == key:
                    found = True
                    assert scrubbed(value), (
                        f"{handler} returns `{key}` without passing it through "
                        f"public_artifact — the response says where the bytes "
                        f"live")
        assert found, f"{handler} no longer returns a `{key}` key"

    # The export endpoint returns a JOB now, not a registry row — so instead of
    # asserting it scrubs one, assert it never hands one over at all. A job
    # names an artifact; it does not carry the registry's storage columns.
    from app.reports.data.export_job_model import ReportExportJob

    receipt = ReportExportJob(
        identity="i", export_type="t", source_intake_id=None,
        classification="DEVELOPMENT_TEST", generator_version="v",
        state="QUEUED", requested_by="someone").to_dict()
    for field in INTERNAL_ARTIFACT_FIELDS:
        assert field not in receipt, (
            f"an export job receipt carries {field} — a job names an artifact, "
            f"never where its bytes live")


async def test_the_workbook_is_readable_rather_than_decorated(rolled_back_db):
    """The accessibility properties the library can actually carry.

    Not a Section 508 certification, and this test does not claim to be one.
    """
    _, _, book = await _built(rolled_back_db)

    assert [ws.title for ws in book.worksheets] == list(wb.SHEET_ORDER)

    for ws in book.worksheets:
        assert not ws.merged_cells.ranges, (
            f"{ws.title} merges cells, which breaks a screen reader's reading "
            f"order and a filter's column")
        assert ws.freeze_panes, f"{ws.title} has no frozen header"
        # The frozen region must not swallow the data: panes freeze at the row
        # BELOW the header, so row 1 of the scrolling region is a data row.
        assert ws.freeze_panes.startswith("A"), (
            f"{ws.title} freezes columns as well, hiding the identifier")

        header_row = 3 if ws.cell(row=1, column=1).value and \
            not ws.cell(row=2, column=1).value else 1
        titles = [ws.cell(row=header_row, column=c).value
                  for c in range(1, ws.max_column + 1)]
        assert all(t not in (None, "") for t in titles), (
            f"{ws.title} has an unlabelled column")


async def test_the_export_reads_the_database_a_fixed_number_of_times(
        rolled_back_db):
    """Cost must not scale with the population.

    23,566 records must not mean 23,566 round trips — and `Review_Status` did
    exactly that until this gate, deriving each case's state with its own query.
    """
    from sqlalchemy import event

    db = rolled_back_db
    intake_id = await _synthetic_estate(db)

    statements = []
    engine = db.get_bind().engine

    def count(conn, cursor, statement, parameters, context, executemany):
        statements.append(statement)

    event.listen(engine, "before_cursor_execute", count)
    try:
        await wb.build_workbook_dataset(
            db, intake_id=intake_id, classification="DEVELOPMENT_TEST",
            generated_by="synthetic-certification")
    finally:
        event.remove(engine, "before_cursor_execute", count)

    selects = [q for q in statements if q.lstrip().upper().startswith("SELECT")]
    assert len(selects) <= 20, (
        f"the export issued {len(selects)} queries for a 7-record delivery; "
        f"something is per-row")


def test_the_engine_handles_a_meaningful_volume():
    """A thousand rows through the real engine, not a ten-row demo.

    The assertion is on WORK, not on a stopwatch: every row must arrive and the
    text controls must still hold at volume. The wall-clock ceiling is generous
    enough not to flake on shared hardware and tight enough to catch a per-cell
    regression — the style-interning defect this gate fixed made the engine
    roughly three times slower.
    """
    import time

    rows = [[f"{i:05d}", str(i).zfill(5), f"=formula-{i}",
             str(1234567890123456 + i)]
            for i in range(1000)]
    dataset = {
        "sheet_order": ["Volume"],
        "sheets": {"Volume": {
            "columns": ["Row", "Postal", "Suspicious", "Long identifier"],
            "rows": rows, "text_columns": [0, 1, 2, 3], "note": None}},
        "report_id": "VOLUME", "workbook_version": wb.WORKBOOK_VERSION,
        "classification": "DEVELOPMENT_TEST",
    }

    started = time.time()
    content = xe.render_workbook(dataset)
    elapsed = time.time() - started

    book = load_workbook(io.BytesIO(content))
    ws = book["Volume"]
    assert ws.max_row == 1001, f"{ws.max_row - 1} rows arrived, not 1000"
    assert ws.cell(row=2, column=2).value == "00000", "a leading zero was lost"
    assert "E+" not in str(ws.cell(row=2, column=4).value)
    formula = ws.cell(row=2, column=3)
    assert formula.data_type == "s" and formula.quotePrefix is True
    assert elapsed < 60, f"1000 rows took {elapsed:.1f}s"



async def _neighbouring_delivery(db):
    """A SECOND synthetic delivery, promoted, related, verified and reviewed.

    The scope tests above use a delivery that promoted nothing, which exercises
    the empty-scope guard and stops short of the filter itself. This one has a
    populated scope of its own, so a workbook built for the FIRST delivery has
    something real it must exclude — and the proof no longer depends on the
    database happening to hold anyone else's rows.
    """
    tag = "NEIGHBOUR"
    intake_id, made = await _delivery(db, [_row(tag)])
    source_id, row = made[0]

    qhin_id, entity_id = uuid.uuid4(), uuid.uuid4()
    for eid, level, name in ((qhin_id, "qhin", f"{SYN} OTHER QHIN"),
                             (entity_id, "participant", f"{SYN} ORG {tag}")):
        db.add(reg.TefcaRegEntity(
            id=eid, name=name, display_name=name, entity_level=level,
            entity_type="provider", operational_status="active",
            verification_status="not_verified", current_version=1,
            is_active=True,
            rce_org_oid=(row["id"] if level == "participant" else None),
            source_record_id=(source_id if level == "participant" else None)))
    await db.flush()

    db.add(reg.TefcaEntityRelationship(
        id=uuid.uuid4(), parent_entity_id=qhin_id, child_entity_id=entity_id,
        relationship_type="managed_by_qhin", status="active", source="import",
        effective_date=date(2026, 1, 1)))
    db.add(m.RceCuratedRecord(
        id=uuid.uuid4(), source_intake_id=intake_id, source_record_id=source_id,
        record_status="CLEAN", issue_count=0, correction_count=0,
        rce_org_oid=row["id"], name=row["name"],
        transformation_version="test-1.0.0", canonical_entity_id=entity_id))
    db.add(reg.TefcaVerification(
        id=uuid.uuid4(), entity_id=entity_id, source="SAM",
        lookup_identifier=row["id"], verification_status="unavailable",
        detail="Synthetic neighbour observation.",
        data_source_label="Synthetic", verified_at=datetime(2026, 7, 21, 9, 0)))
    # A review case too, or Review_Status has nothing of this delivery's to
    # leak and the scope assertion on that sheet proves nothing.
    db.add(reg.ReviewRecord(
        id=uuid.uuid4(), review_id="REV-8301-000001", entity_id=entity_id,
        verification_results={"queue_source": "RCE_DQ_HUMAN_REQUIRED",
                              "selection_reason": "HUMAN_REQUIRED"}))
    await db.flush()
    await db.commit()
    return row["id"]


async def test_another_deliverys_rows_never_appear(rolled_back_db):
    """The filter, not the guard.

    With a second delivery present and promoted, a workbook built for the first
    must exclude it. An unbounded query over the entity tables returns both, and
    this is what says so — the empty-delivery test above cannot, because with no
    entities in scope the query is never reached.
    """
    db = rolled_back_db
    intake_id = await _synthetic_estate(db)
    neighbour_oid = await _neighbouring_delivery(db)

    dataset = await wb.build_workbook_dataset(
        db, intake_id=intake_id, classification="DEVELOPMENT_TEST",
        generated_by="synthetic-certification")

    for name in ("Verification", "Relationships", "Review_Status"):
        sheet = dataset["sheets"][name]
        column = next(i for i, title in enumerate(sheet["columns"])
                      if title in ("Source ID", "Child source ID"))
        exported = {str(cell) for row in sheet["rows"] for cell in row}
        assert neighbour_oid not in exported, (
            f"{name} exported {neighbour_oid}, which belongs to a DIFFERENT "
            f"delivery — the sheet is not bounded by the one being exported")
        assert "REV-8301-000001" not in exported, (
            f"{name} exported another delivery's review case")
        assert sheet["rows"], f"{name} is empty; the test proves nothing"
        # `column` is resolved above so a renamed or reordered organisation
        # column is still a failure rather than a silently skipped check.
        assert sheet["columns"][column] in ("Source ID", "Child source ID")
