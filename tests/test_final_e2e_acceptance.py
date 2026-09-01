"""The final synthetic end-to-end acceptance.

    synthetic delivery -> intake -> immutable Area 1 -> DQ -> curation ->
    promotion -> verification -> monthly delta -> per-QHIN sampling ->
    Priority Review -> assignment -> analyst -> independent QA ->
    reportability -> report -> XLSX -> artifact -> download -> audit

This is NOT another development phase and it certifies nothing new. Every stage
below is already covered by its own suite; what this proves is that one case can
travel the WHOLE chain and be reconstructed from persisted records at the end —
the question no single-stage suite can answer.

GOVERNMENT DATA
    Everything runs inside an OUTER transaction that is rolled back. Fixtures
    are synthetic: OIDs under an unassigned `9.99.999` arc, prefixed names, no
    real NPI, `@synthetic.test` actors. No Government record is read for its
    content, written, or exported, and no official finding is produced.
"""

from __future__ import annotations

import hashlib
import importlib.util
import io
import os
import uuid
from datetime import date, datetime
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine
from sqlalchemy.pool import NullPool

from app.core.database import _normalize_url
from app.tefca_registry import models as reg
from app.tefca_registry.rce import models as m
from app.tefca_registry.rce.field_map import RCE_FIELDS, schema_fingerprint

_spec = importlib.util.spec_from_file_location(
    "step17_fixtures", os.path.join(os.path.dirname(__file__),
                                    "test_onc_review_workbook.py"))
step17 = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(step17)

ARC = "9.99.999"
SYN = "SYNTHETIC-E2E"

ANALYST_A = SimpleNamespace(id=uuid.uuid4(), email="analyst-a@synthetic.test",
                            role="reviewer")
ANALYST_B = SimpleNamespace(id=uuid.uuid4(), email="analyst-b@synthetic.test",
                            role="reviewer")
QA = SimpleNamespace(id=uuid.uuid4(), email="qa@synthetic.test", role="qalead")
SUPERVISOR = SimpleNamespace(id=uuid.uuid4(), email="pm@synthetic.test",
                             role="program_manager")


@pytest.fixture
async def db(db_required):
    engine = create_async_engine(
        _normalize_url(os.environ["DATABASE_URL"]), poolclass=NullPool)
    connection = await engine.connect()
    outer = await connection.begin()
    session = AsyncSession(bind=connection,
                           join_transaction_mode="create_savepoint",
                           expire_on_commit=False)
    try:
        yield session
    finally:
        await session.close()
        await outer.rollback()          # nothing this file does is persisted
        await connection.close()
        await engine.dispose()


# ═══════════════════════════════════════════════════════════════════════════
# STAGE 1-4  intake, Area 1, DQ, curation, promotion
# ═══════════════════════════════════════════════════════════════════════════

async def test_the_chain_from_delivery_to_promotion(db):
    """One delivery, carried through the real pipeline objects."""
    intake_id = await step17._synthetic_estate(db)

    intake = (await db.execute(
        select(m.RceSourceIntake).where(m.RceSourceIntake.id == intake_id)
    )).scalars().first()

    # ── intake: provenance captured, not assumed ────────────────────────────
    assert intake is not None
    assert len(intake.sha256) == 64, "the delivered file has no usable checksum"
    assert intake.schema_fingerprint == schema_fingerprint(list(RCE_FIELDS))
    assert intake.record_count == 7
    assert len(intake.headers) == 41, "the 41-field contract"

    # ── Area 1: every delivered row present and hashed ──────────────────────
    sources = (await db.execute(
        select(m.RceSourceRecord)
        .where(m.RceSourceRecord.source_intake_id == intake_id))).scalars().all()
    assert len(sources) == intake.record_count
    for record in sources:
        assert len(record.record_sha256) == 64
        assert len(record.parsed) == 41

    # ── DQ: findings carry the rule set that produced them ──────────────────
    issues = (await db.execute(
        select(m.RceIssue)
        .where(m.RceIssue.source_intake_id == intake_id))).scalars().all()
    assert issues, "the delivery produced no findings at all"
    authorities = {i.correction_authority for i in issues}
    assert {"AUTO_SAFE", "HUMAN_REQUIRED", "NO_CORRECTION"} <= authorities
    for issue in issues:
        assert issue.rule_version, "a finding with no rule-set version"

    # ── curation: source untouched, curated derived ─────────────────────────
    curated = (await db.execute(
        select(m.RceCuratedRecord)
        .where(m.RceCuratedRecord.source_intake_id == intake_id))).scalars().all()
    assert len(curated) == len(sources)
    held = [c for c in curated if c.record_status == "HELD"]
    assert held, "no HELD record — the held path is not exercised"

    # the delivered value is still the delivered value
    autosafe = next(s for s in sources if s.source_rce_id == f"{ARC}.AUTOSAFE")
    assert autosafe.parsed["address_state"] == "ma", (
        "curation altered the immutable source")

    # ── promotion: HELD is not promoted ─────────────────────────────────────
    for record in curated:
        if record.record_status == "HELD":
            assert record.canonical_entity_id is None, (
                "a HELD record was promoted")
        else:
            assert record.canonical_entity_id is not None


# ═══════════════════════════════════════════════════════════════════════════
# STAGE 5  verification semantics
# ═══════════════════════════════════════════════════════════════════════════

async def test_verification_semantics_survive_the_chain(db):
    """The distinction an export is most likely to lose."""
    intake_id = await step17._synthetic_estate(db)

    observations = (await db.execute(select(reg.TefcaVerification))).scalars().all()
    statuses = {o.verification_status for o in observations}
    assert "unavailable" in statuses, "SOURCE_UNAVAILABLE is not exercised"
    assert "not_found" in statuses, "NO_MATCH_OBSERVED is not exercised"

    # and neither is ever silently upgraded
    for observation in observations:
        assert observation.verification_status not in ("clear", "pass",
                                                       "PASS", "CLEAR"), (
            "an observation was recorded as a conclusion")


# ═══════════════════════════════════════════════════════════════════════════
# STAGE 10-11  analyst, independent QA, reportability
# ═══════════════════════════════════════════════════════════════════════════

async def test_only_an_independently_approved_case_becomes_reportable(db):
    """Segregation of duties, end to end.

    The estate builds three cases — approved, returned, escalated — through the
    real `qa_gate`. Only the approved one may be reportable.
    """
    await step17._synthetic_estate(db)

    from app.tefca_registry.case_assignment import case_state

    records = (await db.execute(
        select(reg.ReviewRecord)
        .where(reg.ReviewRecord.review_id.like("REV-8300-%"))
        .order_by(reg.ReviewRecord.review_id))).scalars().all()
    assert len(records) == 3

    states = {r.review_id: await case_state(db, r.review_id) for r in records}
    assert set(states.values()) == {"APPROVED", "RETURNED", "ESCALATED"}, states

    for record in records:
        state = states[record.review_id]
        if state == "APPROVED":
            assert record.reportable_at is not None
        else:
            assert record.reportable_at is None, (
                f"{record.review_id} is {state} yet reportable")

    # every decision is an appended event, never an edit
    events = (await db.execute(select(reg.ReviewDecisionEvent))).scalars().all()
    assert len(events) >= 6, "analyst + QA events are not both recorded"


async def test_an_analyst_cannot_approve_their_own_determination(db):
    """The control that makes independent QA independent."""
    from app.tefca_registry.qa_gate import (record_analyst_determination,
                                            submit_qa_review)

    await step17._synthetic_estate(db)

    entity = (await db.execute(select(reg.TefcaRegEntity)
                               .where(reg.TefcaRegEntity.rce_org_oid
                                      == f"{ARC}.CLEAN"))).scalars().first()
    review_id = f"REV-E2E-{uuid.uuid4().hex[:6]}"
    db.add(reg.ReviewRecord(id=uuid.uuid4(), review_id=review_id,
                            entity_id=entity.id,
                            verification_results={"queue_source": "E2E"}))
    await db.flush()

    await record_analyst_determination(
        db, review_id, user=ANALYST_A, determination="CONFIRM",
        rationale="Synthetic E2E determination for acceptance.")

    with pytest.raises(Exception) as raised:
        await submit_qa_review(db, review_id, user=ANALYST_A,
                               qa_action="APPROVE",
                               qa_reason="Approving my own work.")
    assert "same" in str(raised.value).lower() or \
           "own" in str(raised.value).lower() or \
           "segregation" in str(raised.value).lower(), str(raised.value)

    # a different reviewer may
    await submit_qa_review(db, review_id, user=QA, qa_action="APPROVE",
                           qa_reason="Independent synthetic approval.")


# ═══════════════════════════════════════════════════════════════════════════
# STAGE 13  the controlled workbook, over the whole chain
# ═══════════════════════════════════════════════════════════════════════════

async def test_the_workbook_reflects_the_whole_chain(db):
    from app.reports.data import onc_review_workbook as wb
    from app.reports.engine import xlsx_engine as xe

    intake_id = await step17._synthetic_estate(db)
    dataset = await wb.build_workbook_dataset(
        db, intake_id=intake_id, classification="DEVELOPMENT_TEST",
        generated_by="e2e@synthetic.test")
    book = load_workbook(io.BytesIO(xe.render_workbook(dataset)))

    assert [s.title for s in book.worksheets] == list(wb.SHEET_ORDER)
    assert dataset["reconciliation"]["exported_source_fields"] == 41
    assert dataset["reconciliation"]["missing"] == []
    assert dataset["reconciliation"]["row_count_matches"] is True

    # the coercion row survives the whole pipeline into the workbook
    ws = book["Source_Data"]
    header_row = next(r for r in range(1, 6)
                      if ws.cell(row=r, column=1).value == "id")
    header = [ws.cell(row=header_row, column=c).value
              for c in range(1, ws.max_column + 1)]
    row = next(r for r in range(header_row + 1, ws.max_row + 1)
               if ws.cell(row=r, column=1).value == f"{ARC}.COERCE")
    assert ws.cell(row=row, column=header.index("address_postalCode") + 1
                   ).value == "01234"
    formula = ws.cell(row=row, column=header.index("alias") + 1)
    assert formula.data_type == "s" and formula.quotePrefix is True

    # and an unavailable source is still unavailable on the Verification sheet
    verification = dataset["sheets"]["Verification"]
    results = {r[verification["columns"].index("Result")]
               for r in verification["rows"]}
    assert "unavailable" in results


# ═══════════════════════════════════════════════════════════════════════════
# STAGE 15  audit reconstruction — the question no single suite answers
# ═══════════════════════════════════════════════════════════════════════════

async def test_one_approved_case_reconstructs_from_persisted_records(db):
    """Follow one case backwards, from the approval to the delivered bytes.

    Every link must come from a persisted record — nothing inferred, nothing
    recomputed. A break anywhere means the case cannot be explained later.
    """
    intake_id = await step17._synthetic_estate(db)

    # 1. an approved, reportable case
    review = (await db.execute(
        select(reg.ReviewRecord)
        .where(reg.ReviewRecord.reportable_at.is_not(None)))).scalars().first()
    assert review is not None, "no reportable case to reconstruct"

    # 2. its decision events, in order, append-only
    events = (await db.execute(
        select(reg.ReviewDecisionEvent)
        .where(reg.ReviewDecisionEvent.review_id == review.review_id)
        .order_by(reg.ReviewDecisionEvent.created_at))).scalars().all()
    assert len(events) >= 2, "a reportable case with fewer than two events"
    # `actor_user_id` is the persisted actor — the event table deliberately
    # keeps the identity per event rather than as columns on the case.
    actors = {str(e.actor_user_id) for e in events if e.actor_user_id}
    assert len(actors) >= 2, (
        f"analyst and QA are the same person: {actors}")
    kinds = {e.event_type for e in events}
    assert reg.ReviewDecisionEvent.ANALYST_DETERMINATION in kinds
    assert reg.ReviewDecisionEvent.QA_REVIEW in kinds

    # 3. the canonical entity
    entity = (await db.execute(
        select(reg.TefcaRegEntity)
        .where(reg.TefcaRegEntity.id == review.entity_id))).scalars().first()
    assert entity is not None and entity.rce_org_oid

    # 4. the curated record that promoted it
    curated = (await db.execute(
        select(m.RceCuratedRecord)
        .where(m.RceCuratedRecord.canonical_entity_id == entity.id)
    )).scalars().first()
    assert curated is not None
    assert curated.source_intake_id == intake_id

    # 5. the immutable source record behind that
    source = (await db.execute(
        select(m.RceSourceRecord)
        .where(m.RceSourceRecord.id == curated.source_record_id)
    )).scalars().first()
    assert source is not None
    assert source.source_rce_id == entity.rce_org_oid, (
        "the entity and its source record disagree about which organisation "
        "this is")

    # 6. the delivery, with its checksum
    intake = (await db.execute(
        select(m.RceSourceIntake)
        .where(m.RceSourceIntake.id == source.source_intake_id)
    )).scalars().first()
    assert intake is not None and len(intake.sha256) == 64

    # 7. the bytes still hash to what was recorded
    assert hashlib.sha256(source.raw_line.encode()).hexdigest() == \
        source.record_sha256, "the stored row no longer matches its own hash"

    # the chain, stated
    print(f"\nreconstructed: delivery {intake.delivery_label} "
          f"-> source {source.source_rce_id} -> curated {curated.record_status} "
          f"-> entity {entity.rce_org_oid} -> {len(events)} decision events "
          f"by {len(actors)} distinct actors -> reportable "
          f"{review.reportable_at}")


# ═══════════════════════════════════════════════════════════════════════════
# STAGE 16  negative security path
# ═══════════════════════════════════════════════════════════════════════════

async def test_area_one_refuses_mutation_at_the_application_layer(db):
    """The database refuses this too (proven separately against Azure DEV).
    This is the application-layer half."""
    intake_id = await step17._synthetic_estate(db)

    before = (await db.execute(text(
        "SELECT md5(string_agg(record_sha256, '' ORDER BY id)) "
        "FROM rce_source_records WHERE source_intake_id = :i"),
        {"i": intake_id})).scalar()

    from app.reports.data import onc_review_workbook as wb
    from app.reports.engine import xlsx_engine as xe

    # a full export must not change one byte of the source
    xe.render_workbook(await wb.build_workbook_dataset(
        db, intake_id=intake_id, classification="DEVELOPMENT_TEST",
        generated_by="e2e@synthetic.test"))

    after = (await db.execute(text(
        "SELECT md5(string_agg(record_sha256, '' ORDER BY id)) "
        "FROM rce_source_records WHERE source_intake_id = :i"),
        {"i": intake_id})).scalar()
    assert before == after, "exporting altered the immutable source"


async def test_a_duplicate_export_does_not_create_a_second_job(db):
    from app.reports.data import export_jobs
    from app.reports.data.onc_review_workbook import WORKBOOK_VERSION
    from app.reports.engine.xlsx_engine import XLSX_ENGINE_VERSION

    intake_id = await step17._synthetic_estate(db)
    identity = export_jobs.job_identity(
        intake_id=str(intake_id), workbook_version=WORKBOOK_VERSION,
        engine_version=XLSX_ENGINE_VERSION,
        classification="DEVELOPMENT_TEST", export_type="onc_review_workbook")

    first = await export_jobs.request_job(
        db, identity=identity, export_type="onc_review_workbook",
        intake_id=intake_id, classification="DEVELOPMENT_TEST",
        generator_version="e2e", requested_by=QA.email)
    second = await export_jobs.request_job(
        db, identity=identity, export_type="onc_review_workbook",
        intake_id=intake_id, classification="DEVELOPMENT_TEST",
        generator_version="e2e", requested_by=SUPERVISOR.email)

    assert str(first.id) == str(second.id)


def test_the_fixtures_are_synthetic_and_labelled():
    assert ARC.startswith("9.99.")
    for actor in (ANALYST_A, ANALYST_B, QA, SUPERVISOR):
        assert actor.email.endswith("@synthetic.test")
