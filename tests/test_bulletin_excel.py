"""The client Excel workbook: shape, branding, and the three-sheet agreement.

The workbook is a deliverable — an FCC contact opens it in Excel — so these
assert the things that make it readable rather than merely well-formed: the
header identifies columns, the summary adds up, and the file survives a round
trip through a real reader.
"""
import io

import pytest
from openpyxl import load_workbook

from app.bulletin_intelligence.excel_export import HEADERS, create_bulletin_excel

BRIEFING = {
    "briefing_id": "fcc_20260731_120000",
    "agency_id": "fcc",
    "briefing_date": "July 31, 2026",
}

EXPECTED_HEADERS = ["#", "Category", "Date", "Relationship", "Title", "Summary",
                    "Source", "Subscription Required", "Relevance", "URL", "Provider"]


class _Art:
    def __init__(self, source, title, summary, url, section, relevance=0.9,
                 paywalled=False, kind="news", published="2026-07-31T09:00:00+00:00",
                 provider="RSS"):
        self.source_name = self.outlet = self.source = source
        self.title, self.summary, self.url = title, summary, url
        self.section, self.topic = section, "x"
        self.relevance_score = relevance
        self.is_paywalled = paywalled
        self.article_type = kind
        self.published_at = published
        self.provider = provider


def _articles():
    return [
        _Art("Radio World", "NAB meets FCC", "One.", "https://e.test/a", "Media & Broadcasting"),
        _Art("Reuters", "Spectrum auction", "Two.", "https://e.test/b", "Wireless & Spectrum",
             relevance=0.5),
        _Art("WaPo", "Ownership rules", "Three.", "https://e.test/c", "Media & Broadcasting",
             paywalled=True),
        _Art("Broadband Breakfast", "Why the FCC is wrong", "Four.", "https://e.test/d",
             "Broadband & Infrastructure", relevance=0.2, kind="opinion"),
    ]


def _roundtrip(wb):
    """Save and reopen — proves a real reader can parse what we produced, which
    a purely in-memory assertion does not."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return load_workbook(buf), buf.getvalue()


def _summary_map(s):
    """Label -> value over the whole Summary sheet, so these tests do not break
    every time a row is inserted."""
    return {s.cell(r, 1).value: s.cell(r, 2).value for r in range(1, s.max_row + 1)}


def test_three_sheets():
    wb, data = _roundtrip(create_bulletin_excel(BRIEFING, _articles(), lambda a: a.section))
    assert wb.sheetnames == ["FCC Daily Bulletin", "Google News Cross-Check", "Summary"]
    assert data[:2] == b"PK", "not a valid xlsx container"


def test_column_order():
    """A through K, exactly as specified — this is what the FCC reads."""
    wb, _ = _roundtrip(create_bulletin_excel(BRIEFING, _articles(), lambda a: a.section))
    ws = wb["FCC Daily Bulletin"]
    assert [ws.cell(1, c).value for c in range(1, len(HEADERS) + 1)] == EXPECTED_HEADERS
    assert ws.cell(1, len(HEADERS) + 1).value is None, "no stray extra column"
    assert ws.max_row == 1 + len(_articles())

    assert ws.cell(2, 1).value == 1
    assert ws.cell(2, 3).value == "2026-07-31"
    assert ws.cell(2, 8).value in ("Yes", "No")
    assert ws.cell(2, 9).value in ("High", "Medium", "Low")
    assert ws.cell(2, 4).value in ("Original", "Follow-up", "Analysis")


def test_summary_sheet_stats():
    wb, _ = _roundtrip(create_bulletin_excel(
        BRIEFING, _articles(), lambda a: a.section,
        qa_report={"google_news_count": 10, "matched": 8}))
    s = wb["Summary"]
    m = _summary_map(s)

    for label in ("Total Articles", "By Category", "By Relevance",
                  "Google News Coverage", "Missing from Google News",
                  "Sources Used", "TOTAL"):
        assert label in m, f"Summary missing {label}"

    assert m["Total Articles"] == 4
    assert m["Sources Used"] == 4
    assert m["Google News Coverage"] == "8/10 (80%)"
    assert m["Media & Broadcasting"] == 2
    # Live formula, not a baked number, so the sheet stays correct if a reader
    # deletes a category row.
    assert str(m["TOTAL"]).startswith("=SUM(")


def test_google_news_sheet():
    """Empty must say so explicitly — a blank sheet reads as 'not run'."""
    wb, _ = _roundtrip(create_bulletin_excel(BRIEFING, _articles(), lambda a: a.section))
    gn = wb["Google News Cross-Check"]
    assert "GOOGLE NEWS CROSS-CHECK" in (gn.cell(1, 1).value or "")
    assert gn.cell(3, 1).value == \
        "All Google News articles matched. No missing stories identified."

    missing = [_Art("AP", "Story only Google News had", "Five.", "https://g.test/9",
                    "Wireless & Spectrum")]
    wb2, _ = _roundtrip(create_bulletin_excel(
        BRIEFING, _articles(), lambda a: a.section, google_news_missing=missing))
    gn2 = wb2["Google News Cross-Check"]
    assert [gn2.cell(2, c).value for c in range(1, len(HEADERS) + 1)] == EXPECTED_HEADERS
    assert gn2.cell(3, 5).value == "Story only Google News had"


def test_rows_are_ordered_by_category_then_source():
    wb, _ = _roundtrip(create_bulletin_excel(BRIEFING, _articles(), lambda a: a.section))
    ws = wb["FCC Daily Bulletin"]
    got = [(ws.cell(r, 2).value, ws.cell(r, 7).value) for r in range(2, ws.max_row + 1)]
    assert got == sorted(got, key=lambda t: (t[0].lower(), t[1].lower()))


def test_reader_affordances_present():
    """Frozen header, auto-filter and clickable URLs are what make 180 rows
    usable; without them the sheet is technically correct and unreadable."""
    wb, _ = _roundtrip(create_bulletin_excel(BRIEFING, _articles(), lambda a: a.section))
    ws = wb["FCC Daily Bulletin"]
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref.startswith("A1:K")
    assert ws.cell(2, 10).hyperlink is not None, "URL column must be clickable"
    assert ws.cell(2, 5).alignment.wrap_text and ws.cell(2, 6).alignment.wrap_text


def test_title_tags_applied_and_kept_out_of_summaries():
    """Tags belong in the title — the summary prompt forbids them in the summary."""
    wb, _ = _roundtrip(create_bulletin_excel(BRIEFING, _articles(), lambda a: a.section))
    ws = wb["FCC Daily Bulletin"]
    titles = " | ".join(str(ws.cell(r, 5).value or "") for r in range(2, ws.max_row + 1))
    summaries = " ".join(str(ws.cell(r, 6).value or "") for r in range(2, ws.max_row + 1))
    assert "[Subscription Required]" in titles
    assert "[Opinion]" in titles
    assert "[Opinion]" not in summaries
    assert "[Subscription Required]" not in summaries


def test_control_characters_do_not_break_the_workbook():
    """Summaries are scraped from third-party feeds; one stray control byte must
    not cost the whole download."""
    arts = _articles()
    arts[0].summary = "Bad\x07byte\x0bhere"
    wb, _ = _roundtrip(create_bulletin_excel(BRIEFING, arts, lambda a: a.section))
    ws = wb["FCC Daily Bulletin"]
    joined = "".join(str(ws.cell(r, 6).value or "") for r in range(2, ws.max_row + 1))
    assert "\x07" not in joined and "\x0b" not in joined


def test_empty_briefing_still_produces_a_valid_file():
    wb, data = _roundtrip(create_bulletin_excel(BRIEFING, [], lambda a: "General"))
    assert data[:2] == b"PK"
    assert wb.sheetnames == ["FCC Daily Bulletin", "Google News Cross-Check", "Summary"]
    assert wb["FCC Daily Bulletin"].max_row == 1  # header only
    assert _summary_map(wb["Summary"])["Total Articles"] == 0


# ── HTML must never reach a cell ─────────────────────────────────────────────

def test_no_html_in_any_cell():
    """The reported defect: raw markup rendered into the Summary column."""
    arts = _articles()
    arts[0].summary = "<p>The Commission voted <b>today</b>.</p>"
    arts[0].title = "<span>FCC acts</span> on spectrum"
    arts[1].summary = "<div>Fined &amp;8,000, per the <a href='#'>order</a>.</div>"
    arts[1].source_name = arts[1].outlet = "AT&amp;T News"

    wb, _ = _roundtrip(create_bulletin_excel(BRIEFING, arts, lambda a: a.section))
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows():
            for cell in row:
                text = str(cell.value or "")
                assert "<" not in text and ">" not in text, \
                    f"HTML leaked into {name}!{cell.coordinate}: {text[:80]!r}"
                assert "&amp;" not in text and "&lt;" not in text, \
                    f"undecoded entity in {name}!{cell.coordinate}: {text[:80]!r}"

    ws = wb["FCC Daily Bulletin"]
    joined = " ".join(str(ws.cell(r, 6).value or "") for r in range(2, ws.max_row + 1))
    assert "The Commission voted today." in joined, "spacing before punctuation not closed"


def test_qa_mode_is_the_client_sheet_plus_l_to_p():
    """'Same as Button 1 plus extras' only holds if A-K is byte-identical."""
    from app.bulletin_intelligence.excel_export import QA_HEADERS
    arts = _articles()
    client, _ = _roundtrip(create_bulletin_excel(BRIEFING, arts, lambda a: a.section))
    qa, _ = _roundtrip(create_bulletin_excel(
        BRIEFING, arts, lambda a: a.section,
        qa=True, qa_extras=lambda a: ["OK", "No", "", 80, "YES"]))

    cw, qw = client["FCC Daily Bulletin"], qa["FCC Daily Bulletin"]
    assert [cw.cell(1, c).value for c in range(1, 12)] == \
           [qw.cell(1, c).value for c in range(1, 12)]
    assert [qw.cell(1, c).value for c in range(12, 17)] == QA_HEADERS
    assert cw.max_column == 11, "client sheet must not carry QA columns"
    assert qw.max_column == 16
    assert qw.cell(2, 12).value == "OK"
    assert qw.cell(2, 16).value == "YES"
